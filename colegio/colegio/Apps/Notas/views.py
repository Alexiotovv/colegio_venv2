from http.client import responses
import json
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.views.generic import ListView, DeleteView
from colegio.Apps.Alumno.models import Alumno
from colegio.Apps.Curso.models import Curso

from colegio.Apps.Notas.models import Notas,NotasComp
from colegio.Apps.Notas.models import SettingNotas

from colegio.Apps.Matricula.models import Matricula
from colegio.Apps.AnoAcademico.models import AnoAcademico
from colegio.Apps.Docente.models import Docente
from colegio.Apps.DocenteCurso.models import DocenteCurso
from colegio.Apps.PeriodoAcademico.models import PAcademico
from colegio.Apps.TempDatos.models import TempDatos
from colegio.Apps.Competencias.models import CompetenciaCurso,Competencias

from django.db import connection ##permite conectar directamente a la base de datos
from openpyxl import Workbook,load_workbook

from datetime import *
	
from colegio.Apps.Notas.forms import NotasForm
from colegio.Apps.Alumno.forms import AlumnoForm
from colegio.Apps.Docente.forms import DocenteForm
from colegio.Apps.Curso.forms import CursoForm
from colegio.Apps.PeriodoAcademico.forms import PAcademicoForm
from colegio.Apps.AnoAcademico.forms import AnoAcademicoForm
from django.contrib.auth.models import User
from django.http import JsonResponse

def ObtenerCompetenciasBimPorAula(request,id):
	competencias=list(CompetenciaCurso.objects.filter(Curso_id=id).values('Competencias__nombre_competencia'))
	return JsonResponse(competencias,safe=False)

def ObtenerCursosBimPorAula(request,nivel):
	cursos=list(Curso.objects.filter(Nivel=nivel,Tipo='CURSO').values())
	return JsonResponse(cursos,safe=False)

def AlumnosBimPorAula(request):
	if request.method=='POST':
		
		ano=request.POST.get('AnoAcademico')
		# paca=request.method.POST.get('PeriodoAcademico')
		gradonivel=request.POST.get('GradoNivel')
		seccion=request.POST.get('Seccion')
		alum=list(Matricula.objects.filter(
			AnoAcademico__id=ano,
			# PAcademico__id=paca,
			Grado=gradonivel,
			Seccion=seccion,
			Alumno__Estado='A').values('id','Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres'))
		return JsonResponse(alum,safe=False)

def RegistroBimNotasPorAula(request):
	paca=PAcademico.objects.all().values('id','Nombre','Status')
	contexto={'paca':paca}
	return render(request, 'Notas/RegistroBimNotasPorAula.html',contexto)

def ActualizarAvanceNotasPorAlumno(request):
	if request.method=='POST':
		matricula=request.POST.get('Matriculae')
		periodo=request.POST.get('PAcademicoe')
		curso=request.POST.get('Cursose')
		data=NotasComp.objects.filter(Matricula_id=matricula,PAcademico_id=periodo,Curso_id=curso)
		for dd in data:
			# print(request.POST.get("nota"+str(dd.id)))
			obj = NotasComp.objects.get(id=request.POST.get(str(dd.id)))
			obj.Nota = request.POST.get("nota"+str(dd.id))
			obj.save()

		data={'ok':'ok'}
		return JsonResponse(data)

def  ObtenerAvanceExistentesEditar(request):
	if request.method=='POST':
		matricula=request.POST.get('Matriculae')
		periodo=request.POST.get('PAcademicoe')
		data=list(NotasComp.objects.filter(Matricula_id=matricula,PAcademico_id=periodo).values())
		return JsonResponse(data,safe=False)

def  ObtenerAvanceExistentes(request):
	if request.method=='POST':
		matricula=request.POST.get('Matricula')
		periodo=request.POST.get('PAcademico')
		data=list(NotasComp.objects.filter(Matricula_id=matricula,PAcademico_id=periodo).values())
		return JsonResponse(data,safe=False)

def GuardaAvanceNotasPorAlumno(request):
	if request.method=='POST':
		idCurso=request.POST.get('Cursos')
		idDocente=Docente.objects.get(User__id=request.user.id)#Toma de Consola
		idMatricula=request.POST.get('Matricula')
		idPacademico=request.POST.get('PAcademico')		
		compes=Competencias.objects.filter(competenciacurso__Curso_id=idCurso).order_by('Orden')
		for cc in compes:
			obj = NotasComp()
			obj.Competencias_id=cc.id
			obj.Curso_id=idCurso
			obj.Docente_id=idDocente.id
			obj.Matricula_id=idMatricula
			obj.PAcademico_id=idPacademico
			obj.Nota=request.POST.get(str(cc.id))
			obj.save()
		data={'ok':'ok'}
		return JsonResponse(data)
		
def ObtenerCompetenciasEditar(request):
	if request.method=='POST':
		matricula=request.POST.get('Matriculae')
		periodo=request.POST.get('PAcademicoe')
		curso=request.POST.get('Cursose')
		data=list(NotasComp.objects.filter(Matricula_id=matricula,PAcademico_id=periodo,Curso_id=curso).order_by('Competencias__Orden').values('id','Nota','Competencias__nombre_competencia'))
		return JsonResponse(data,safe=False)

def ObtenerCompetencias(request,idCurso):
	data=list(Competencias.objects.filter(competenciacurso__Curso_id=idCurso).order_by('Orden').values())
	return JsonResponse(data,safe=False)

def BuscarCursoNivel(request, nivel):
	cursos_nivel=list(Curso.objects.filter(Tipo__in=['CURSO','INASISTENCIAS','ACTITUDINAL','APRECIACIÓN DEL TUTOR'],Nivel=nivel).order_by('Orden').values())#CURSOS
	return JsonResponse(cursos_nivel,safe=False)

def RegistroPorAlumno(request):
	#NIVEL SE VA DIRECTO EN EL FRONT
	aac=AnoAcademico.objects.all().order_by('-id')#AÑO
	pac=PAcademico.objects.filter(Status='Activo').order_by('id')#PERIODO ACADÉMICO
								#SECCIÓN ESTÁ EN EL FRONT
	cur=Curso.objects.filter(Tipo__in=['CURSO','INASISTENCIAS','ACTITUDINAL','APRECIACIÓN DEL TUTOR'],Nivel='PRIM').order_by('Orden')#CURSOS
								#GRADO NIVEL ESTÁ EN EL FRONT
	lista_periodos=PAcademico.objects.all().order_by('-id')#todos los periodos
	ano_actual=datetime.now().year
	alu=Matricula.objects.filter(AnoAcademico__Ano=ano_actual)
	contexto={'aac':aac,'pac':pac,'cur':cur,'alu':alu,'lista_periodos':lista_periodos}
	return render(request,'Notas/RegistroBimNotasPorAlumno.html',contexto)
	
def ConsolidadoNotas(request):
	aac=AnoAcademico.objects.all().order_by('-id')
	pac=PAcademico.objects.all().order_by('id')
		
	if	request.method=='POST':
		ano=request.POST.get('ano')
		paca=request.POST.get('pacademico')#Determina la Vista que se seleccionará
		gradonivel=request.POST.get('Grado')
		seccion=request.POST.get('Seccion')
		final=request.POST.get('final')
		nivel=str(gradonivel)[1:len(gradonivel)]
		for p in pac:
			if str(p.id)==paca:
				periodo=p.Nombre
		######con matriculas se obtiene todos los alumnos y se guarda en mtriculas de ese grado filtrado por ano,gradonivel, seccion
		####tambien se obtiene consolidado_notas de NotasComp
		matriculas=Matricula.objects.filter(AnoAcademico__Ano=ano,Grado=gradonivel,Seccion=seccion).order_by('Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres')
		consolidado_notas=NotasComp.objects.filter(Matricula__AnoAcademico__Ano=ano,
		Matricula__Grado=gradonivel,Matricula__Seccion=seccion,PAcademico=paca)
		cursos= CompetenciaCurso.objects.filter(Curso__Nivel=nivel).exclude(Curso__Tipo='--').order_by('Curso__Orden','Competencias__Orden')
		
		###Mi funcion estrella que hace la granputa colocacion en el excel
		if final == '1':
			InsertaNotasEnExcel(consolidado_notas,matriculas,nivel,gradonivel,seccion,periodo,cursos)#funcion que coloca las notas en el excel se encuentra mas abajo
		else:
			InsertaNotasEnExcelPromedio(consolidado_notas,matriculas,nivel,gradonivel,seccion,periodo,cursos)#funcion que coloca las notas en el excel se encuentra mas abajo
		
		############################################3
		if str(nivel).find("PRIM")!= -1:
			if final == '1':
				return redirect("https://colcoopcv.com/static/files/PLANTILLA_LIBRETA_PRIMARIA.xlsx")
				#return redirect("/static/files/PLANTILLA_LIBRETA_PRIMARIA.xlsx")
			else:
				return redirect("https://colcoopcv.com/static/files/PLANTILLA_LIBRETA_PRIMARIA_R.xlsx")
				#return redirect("/static/files/PLANTILLA_LIBRETA_PRIMARIA_R.xlsx")
		else:
			if final=='1':
				return redirect("https://colcoopcv.com/static/files/PLANTILLA_LIBRETA_SECUNDARIA.xlsx")
				#return redirect("/static/files/PLANTILLA_LIBRETA_SECUNDARIA.xlsx")
			else:
				return redirect("https://colcoopcv.com/static/files/PLANTILLA_LIBRETA_SECUNDARIA_R.xlsx")
				#return redirect("/static/files/PLANTILLA_LIBRETA_SECUNDARIA_R.xlsx")
	else:
		contexto={'aac':aac,'pac':pac}
		return render(request,'otras_opciones/imprimir_consolidado_libretas.html',contexto)

def NotasNuevoComp(request):
	persona=TempDatos.objects.filter(User=request.user.id)
	if persona.exists():
		TempDatos.objects.filter(User=request.user.id).delete()

	ano = AnoAcademico.objects.get(Ano=datetime.now().year)#Año
	paca = PAcademico.objects.get(FechaInicio__lt=datetime.now(),FechaFinal__gt=datetime.now())#Bimestre
	doce = Docente.objects.get(User__id=request.user.id)#Profesor

	curso_list = DocenteCurso.objects.filter(Docente__User__id=request.user.id)#para refencia en GradoNivel plantilla Seccion
	
	docente = Docente.objects.get(User__id=request.user.id)
	grados_list = str(docente.GradoNivel).split()#Extrae los grados de la cadena y los pone en lista
	secciones_list = str(docente.Seccion).split()#Extrae las secciones de la cadena y los pone en lista
	######################
	#este contexto es para el ELSE
	contexto2 = {'doce':doce,'paca':paca,'ano':ano,'curso_list':curso_list,'grados_list':grados_list,'secciones_list':secciones_list} #'doce':doce}#'matri':matri
	if request.method=='POST':
		temp_datos = TempDatos()
		
		usu=User()
		usu.id=request.user.id
		doce.User=usu
		
		temp_datos.User = usu#nuevo
		temp_datos.idCurso = request.POST.get("curso")
		temp_datos.grado = request.POST.get("grados")
		temp_datos.seccion = request.POST.get("secciones")
		temp_datos.save()

		return redirect('app_nota_nuevo_save_comp')#aqui guarda los avances de notas secundaria
	else:
		return render(request,'Notas/create_notas_comp.html',contexto2)

def NotasNuevoSaveComp(request):	
	ano = AnoAcademico.objects.get(Ano=datetime.now().year)#Año
	paca = PAcademico.objects.get(FechaInicio__lt=datetime.now(),FechaFinal__gt=datetime.now())#Bimestre
	doce = Docente.objects.get(User__id=request.user.id)#Profesor
	
	usu=User()
	usu.id=request.user.id
	doce.User=usu

	#td = AvanceTempDatos.objects.get(User=usu) #funcionaba
	td_comp=TempDatos.objects.get(User=usu)

	#compe = Competencias.objects.get(id=td_comp.idCurso)
	curso = Curso.objects.get(id=td_comp.idCurso)

	grado =  td_comp.grado
	seccion = td_comp.seccion	
	notas = NotasComp.objects.filter(Curso=td_comp.idCurso, Matricula__Seccion=td_comp.seccion, Matricula__Grado=td_comp.grado, Matricula__AnoAcademico=ano.id, PAcademico__id=paca.id)
	#notas = AvanceNotas.objects.filter(Curso__id=td.idCurso, Matricula__Seccion=td.seccion, Matricula__Grado=td.grado, Matricula__AnoAcademico=ano.id, PAcademico__id=paca.id)
	#Encontrando las competencias del Curso
	lista_compe=CompetenciaCurso.objects.filter(Curso__id=curso.id).order_by('Competencias__Orden')
	######################################################################
	nivel=str(grado)[1:len(grado)]
	num_compe=0
	for num in lista_compe:
		num_compe+=1

	if notas:
		alumnos = ''		
	else:
		alumnos = Matricula.objects.filter(Grado=td_comp.grado,Seccion=td_comp.seccion,AnoAcademico=ano.id).order_by('Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres')
	
	contexto={'grado':grado,'seccion':seccion,'curso':curso,'ano':ano,'paca':paca,'doce':doce,'alumnos':alumnos,'lista_compe':lista_compe,'num_compe':num_compe,'nivel':nivel}
	
	if request.method=='POST':
		for alu in alumnos:
			for cc in lista_compe:
				notacomp = NotasComp()

				cur = Curso()
				cur.id = td_comp.idCurso
				notacomp.Curso = cur

				comp = Competencias()
				comp.id = cc.Competencias.id
				notacomp.Competencias = comp

				pac = PAcademico()
				pac.id = paca.id
				notacomp.PAcademico = pac

				doc = Docente()
				doc.id = doce.id
				notacomp.Docente = doc
		
				mat = Matricula()
				mat.id = alu.id
				notacomp.Matricula = mat

				#es alu.id porque el inputtextNota lleva el nombre del id de matricula
				evalu=alu.id #obtener el id del alumno
				evcomp=cc.Competencias.id##obtener el id de la competencia para combinarlo
				
				nota_eva = request.POST.get(str(evalu)+str(evcomp))
				if nota_eva=='':
					nota_eva='--'#ante era competencias no desarrollada
				if nota_eva=='0' or nota_eva=='1' or nota_eva=='2' or nota_eva=='3' or nota_eva=='4' or nota_eva=='5' or nota_eva=='6' or nota_eva=='7' or nota_eva=='8' or nota_eva=='9' :
					nota_eva='0'+str(nota_eva)
				notacomp.Nota = str(nota_eva).upper()
				notacomp.save()
		grabo='registrado'
		
		notita = NotasComp.objects.filter(Docente__User__id=request.user.id, Curso__id=td_comp.idCurso, Matricula__Grado=td_comp.grado, Matricula__Seccion=td_comp.seccion,Matricula__AnoAcademico=ano.id,PAcademico__id=paca.id).order_by('Matricula__Alumno__ApellidoPaterno','Matricula__Alumno__ApellidoMaterno','Matricula__Alumno__Nombres')
		context = {'grabo':grabo,'curso':curso,'grado':grado,'seccion':seccion,'notita':notita}

		persona=TempDatos.objects.filter(User=request.user.id)
		if persona.exists():
			TempDatos.objects.filter(User=request.user.id).delete()
		return render(request,'Notas/listar_notas.html', context)
	else:
		return render(request,'Notas/create_notas_save_comp.html',contexto)

def NotasNuevoBimestre(request):
	#primero debe borrar los datos de la tabla del usuario actual
	#TempDatos.objects.filter(idusuario=request.user.id).delete()
	persona=TempDatos.objects.filter(User=request.user.id)
	if persona.exists():
		TempDatos.objects.filter(User=request.user.id).delete()

	ano = AnoAcademico.objects.get(Ano=datetime.now().year)#Año
	doce = Docente.objects.get(User__id=request.user.id)#Profesor

	paca_list = PAcademico.objects.all()#Todos los Bimestres
	curso_list = DocenteCurso.objects.filter(Docente__User__id=request.user.id)#para refencia en GradoNivel plantilla Seccion

	docente = Docente.objects.get(User__id=request.user.id)
	grados_list = str(docente.GradoNivel).split()#Extrae los grados de la cadena y los pone en lista
	secciones_list = str(docente.Seccion).split()#Extrae las secciones de la cadena y los pone en lista
	######################
	#este contexto es para el ELSE
	contexto2 = {'doce':doce,'paca_list':paca_list,'ano':ano,'curso_list':curso_list,'grados_list':grados_list,'secciones_list':secciones_list} #'doce':doce}#'matri':matri


	if request.method=='POST':
		temp_datos = TempDatos()
		
		usu=User()
		usu.id=request.user.id
		doce.User=usu
		
		temp_datos.User = usu#nuevo
		temp_datos.idCurso = request.POST.get("curso")
		temp_datos.grado = request.POST.get("grados")
		temp_datos.seccion = request.POST.get("secciones")
		temp_datos.idPAcademico = request.POST.get("bimestres")
		temp_datos.save()
		return redirect('app_nota_nuevo_save_bimestre')
	else:
		return render(request,'Notas/create_notas_bimestre.html',contexto2)

def NotasNuevoSave(request):	
	ano = AnoAcademico.objects.get(Ano=datetime.now().year)#Año
	paca = PAcademico.objects.get(FechaInicio__lt=datetime.now(),FechaFinal__gt=datetime.now())#Bimestre
	doce = Docente.objects.get(User__id=request.user.id)#Profesor
	
	#tempd = TempDatos.objects.latest('id') #funcionaba
	usu=User()
	usu.id=request.user.id
	doce.User=usu

	td = TempDatos.objects.get(User=usu) #funcionaba
	curso = Curso.objects.get(id=td.idCurso)
	grado =  td.grado
	seccion = td.seccion#Aqui es el truco para filtrar que Alumnos tienen Nota o No
	notas = Notas.objects.filter(Curso__id=td.idCurso, Matricula__Seccion=td.seccion, Matricula__Grado=td.grado, Matricula__AnoAcademico=ano.id,PAcademico__id=paca.id)

	if notas:
		alumnos = ''		
	else:
		alumnos = Matricula.objects.filter(Grado=td.grado,Seccion=td.seccion,AnoAcademico=ano.id).order_by('Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres')
	
	contexto={'grado':grado,'seccion':seccion,'curso':curso,'ano':ano,'paca':paca,'doce':doce,'alumnos':alumnos}
	
	if request.method=='POST':

		for alu in alumnos:
			nota = Notas()
			cur = Curso()
			cur.id = td.idCurso
			nota.Curso = cur

			pac = PAcademico()
			pac.id = paca.id
			nota.PAcademico = pac

			doc = Docente()
			doc.id = doce.id
			nota.Docente = doc
	
			mat = Matricula()
			mat.id = alu.id
			nota.Matricula = mat

			#es alu.id porque el inputtextNota lleva el nombre del id de matricula
			evalu=alu.id
			nota_eva = request.POST.get(str(evalu))
			nota.Nota = nota_eva.upper()
			nota.save()
			
		grabo='registrado'
		
		notita = Notas.objects.filter(Docente__User__id=request.user.id, Curso__id=td.idCurso, Matricula__Grado=td.grado, Matricula__Seccion=td.seccion,Matricula__AnoAcademico=ano.id,PAcademico__id=paca.id).order_by('Matricula__Alumno__ApellidoPaterno','Matricula__Alumno__ApellidoMaterno','Matricula__Alumno__Nombres')
		context = {'grabo':grabo,'curso':curso,'grado':grado,'seccion':seccion,'notita':notita}

		persona = TempDatos.objects.filter(User=request.user.id)
		if persona.exists():
			TempDatos.objects.filter(User=request.user.id).delete()
		return render(request,'Notas/listar_notas.html',context)
	else:
		return render(request,'Notas/create_notas_save.html',contexto)

def NotasNuevoSaveBimestre(request):	
	ano = AnoAcademico.objects.get(Ano=datetime.now().year)#Año
	doce = Docente.objects.get(User__id=request.user.id)#Profesor
	
	#tempd = TempDatos.objects.latest('id') #funcionaba
	usu=User()
	usu.id=request.user.id
	doce.User=usu

	td = TempDatos.objects.get(User=usu) #funcionaba
	
	curso = Curso.objects.get(id=td.idCurso)
	paca = PAcademico.objects.get(id=td.idPAcademico)
	
	grado =  td.grado
	seccion = td.seccion#Aqui es el truco para filtrar que Alumnos tienen Nota o No
	notas = Notas.objects.filter(Curso__id=td.idCurso, Matricula__Seccion=td.seccion, Matricula__Grado=td.grado, Matricula__AnoAcademico=ano.id,PAcademico__id=td.idPAcademico)

	if notas:#Si existen Alumnos entonces devuelve alumnos vacio
		alumnos = ''		
	else:
		alumnos = Matricula.objects.filter(Grado=td.grado,Seccion=td.seccion,AnoAcademico=ano.id).order_by('Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres')
	
	contexto={'grado':grado,'seccion':seccion,'curso':curso,'ano':ano,'paca':paca,'doce':doce,'alumnos':alumnos}
	
	if request.method=='POST':

		for alu in alumnos:
			nota = Notas()
			cur = Curso()
			cur.id = td.idCurso
			nota.Curso = cur

			pac = PAcademico()
			pac.id = td.idPAcademico
			nota.PAcademico = pac

			doc = Docente()
			doc.id = doce.id
			nota.Docente = doc
	
			mat = Matricula()
			mat.id = alu.id
			nota.Matricula = mat

			#es alu.id porque el inputtextNota lleva el nombre del id de matricula
			evalu=alu.id
			nota_eva = request.POST.get(str(evalu))
			nota.Nota = nota_eva.upper()
			nota.save()
			
		grabo='registrado'
		
		notita = Notas.objects.filter(Docente__User__id=request.user.id, Curso__id=td.idCurso, Matricula__Grado=td.grado, Matricula__Seccion=td.seccion,Matricula__AnoAcademico=ano.id,PAcademico__id=paca.id).order_by('Matricula__Alumno__ApellidoPaterno','Matricula__Alumno__ApellidoMaterno','Matricula__Alumno__Nombres')
		context = {'grabo':grabo,'curso':curso,'grado':grado,'seccion':seccion,'notita':notita}

		persona = TempDatos.objects.filter(User=request.user.id)
		if persona.exists():
			TempDatos.objects.filter(User=request.user.id).delete()
		return render(request,'Notas/listar_notas.html',context)
	else:
		return render(request,'Notas/create_notas_save_bimestre.html',contexto)

def NotasNuevoSaveUno(request):
	ano = AnoAcademico.objects.get(Ano=datetime.now().year)
	paca = PAcademico.objects.filter()
	doce = Docente.objects.get(User__id=request.user.id)#Profesor
	curso = Curso.objects.all()
	matri = Matricula.objects.filter(AnoAcademico=ano.id)
	contexto={'curso':curso,'matri':matri,'ano':ano,'paca':paca}
	
	if request.method=='POST':
		notas = Notas()

		pac = PAcademico()
		pac.id = request.POST.get("PAcademico")
		notas.PAcademico = pac

		doc = Docente()
		doc.id = doce.id
		notas.Docente = doc

		cur = Curso()
		cur.id = request.POST.get("Curso")
		notas.Curso = cur
		
		mat = Matricula()
		mat.id = request.POST.get("Alumno")
		notas.Matricula = mat

		notas.Nota = request.POST.get("Nota")
		notas.save()
		return redirect('app_listar_notas')
	else:
		return render(request,'Notas/create_notas_save_uno.html',contexto)

def NotasEdit(request, id_notas):
	nota = NotasComp.objects.get(id=id_notas)
	if request.method =='GET':
		form = NotasForm(instance=nota)
	else:
		form = NotasForm(request.POST,instance=nota)
		if form.is_valid():
			form.save()
		return redirect('app_listar_notas')#tambien se puede usar return redirect('app_listar_notas')
	contexto = {'nota':nota,'form':form}
	return render(request, 'Notas/editar_notas.html',contexto)

def ListaNotas(request):###corregir
	pac = PAcademico.objects.filter(Status='Activo')

	ingnotas=SettingNotas.objects.get(id=1)
	ingavances=SettingNotas.objects.get(id=2)
	ingnotas=EstadoValor(ingnotas.Valor)
	ingavances=EstadoValor(ingavances.Valor)
	
	if ingnotas=='checked=""':
		chk_bimestre='on'
	else:
		chk_bimestre=''

	if ingavances=='checked=""':
		chk_avance='on'
	else:
		chk_avance=''
	
	contexto={'pac':pac,'ingnotas':ingnotas, 'ingavances':ingavances,'chk_bimestre':chk_bimestre,'chk_avance':chk_avance}

	return render(request, 'Notas/listar_notas.html',contexto)#antes contexto

def OpcionNotas(request):
	#Modificado el 01 Abril
	if request.method=='POST':
		IngresoNotas=request.POST.get('ingNotas')
		IngresoAvances=request.POST.get('ingAvances')

		if IngresoNotas=='on':
			IngresoNotas=True
		else:
			IngresoNotas=False
		
		if IngresoAvances=='on':
			IngresoAvances=True
		else:
			IngresoAvances=False
			
		SettingNotas.objects.filter(id=1).update(Valor=IngresoNotas)
		SettingNotas.objects.filter(id=2).update(Valor=IngresoAvances)
		IngresoNotas=EstadoValor(IngresoNotas)
		IngresoAvances=EstadoValor(IngresoAvances)
		
		contexto2={'ingnotas':IngresoNotas,'ingavances':IngresoAvances}
		return render(request,'Notas/opcion_notas.html',contexto2)
	else:
		ingnotas=SettingNotas.objects.get(id=1)
		ingavances=SettingNotas.objects.get(id=2)
		ingnotas=EstadoValor(ingnotas.Valor)
		ingavances=EstadoValor(ingavances.Valor)
		
		if ingnotas=='checked=""':
			chk_bimestre='on'
		else:
			chk_bimestre=''

		if ingavances=='checked=""':
			chk_avance='on'
		else:
			chk_avance=''

		contexto={'ingnotas':ingnotas, 'ingavances':ingavances,'chk_bimestre':chk_bimestre,'chk_avance':chk_avance}
		return render(request,'Notas/opcion_notas.html',contexto)
class NotasDelete(DeleteView):
	model=Notas
	template_name = 'Notas/delete_notas.html'
	success_url ='/notas/listar/' 

def DeleteNotasxCurso(request):
	Oano=AnoAcademico.objects.all().order_by('-id')
	Opac=PAcademico.objects.all().order_by('id')
	Ocur=Curso.objects.all().order_by('Nivel','Orden')
	Omsje=''
	contexto={'ano':Oano,'pac':Opac,'cur':Ocur,'msje':Omsje}
	if request.method=='POST':
		#Eliminar las Notas x Curso que se envia a traves del POST
		cur=Curso()
		ano=AnoAcademico()
		pac=PAcademico()

		ano.id=request.POST.get("Ano")
		pac.id=request.POST.get("PAcademico")
		cur.id=request.POST.get("Curso")
		gra=request.POST.get("Grado")
		sec=request.POST.get("Seccion")

		NotasComp.objects.filter(Matricula__AnoAcademico=ano.id, PAcademico__id=pac.id,Curso=cur.id, Matricula__Grado=gra,Matricula__Seccion=sec).delete()
		
		Omsje='eliminado'
		contexto={'ano':Oano,'pac':Opac,'cur':Ocur,'msje':Omsje}
		return render(request,'Notas/delete_notas_xcurso.html',contexto)
	else:
		msje='-'
		return render(request,'Notas/delete_notas_xcurso.html',contexto)
	
def EstadoValor(obj):
	if obj:#Si es True
		devuelve='checked=""'
	else:
		devuelve=''

	return (devuelve)


def InsertaNotasEnExcel(consolidado_notas,matriculas,nivel,gradonivel,seccion,periodo,cursos):
	#LIMPIANDO TODAS LAS CELDAS################
	if str(nivel)=='SEC':		
		Ruta = "/var/www/vhosts/colegio_venv/colegio/static/files/PLANTILLA_LIBRETA_SECUNDARIA.xlsx"
		#Ruta = "static/files/PLANTILLA_LIBRETA_SECUNDARIA.xlsx"

		Libro =load_workbook(Ruta)			
		Hoja1 = Libro.active
		fila=6
		col=3
		Hoja1["C3"]=""#Limpiando PeriodoAcademico Grado y Seccion
		for filas in range(7,65):#LIMPIANDO_FILAS
			Hoja1["B"+str(filas)]=""
			Hoja1.cell(row=4,column=filas-3,value="")
			Hoja1.cell(row=5,column=filas-3,value="")
			Hoja1.cell(row=6,column=filas-3,value="")
			for columnas in range(2,65):#LIMPIANDO_COLUMNAS
				Hoja1.cell(row=filas,column=columnas,value="")
			
		Hoja1["C3"]=periodo+" - "+gradonivel+" - "+seccion
		
		#poniendo los cursos y las competencias en los encabezados del excel
		cursocambio=''
		compe=0
		for cur in cursos:
			col+=1
			Hoja1.cell(row=4,column=col,value=cur.Curso.Nombre)
			Hoja1.cell(row=5,column=col,value=cur.Competencias.nombre_competencia)
			if cursocambio!=cur.Curso.Nombre:
				cursocambio=cur.Curso.Nombre
				compe=1
			else:
				compe+=1
			Hoja1.cell(row=6,column=col,value='C'+str(compe))
		
		colnota=3
		for mat in matriculas:#recorre cada fila
			fila+=1
			colnota+=1
			Hoja1["B"+str(fila)] = mat.Alumno.DNI
			Hoja1["C"+str(fila)] = mat.Alumno.ApellidoPaterno+" "+mat.Alumno.ApellidoMaterno+" "+mat.Alumno.Nombres
			for obj in consolidado_notas:
				if mat.id == obj.Matricula.id:
					for colcur in range(4,col+1):
						if obj.Curso.Nombre==Hoja1.cell(row=4,column=colcur).value and obj.Competencias.nombre_competencia==Hoja1.cell(row=5,column=colcur).value:##para saber si cambia curso
							Hoja1.cell(row=fila,column=colcur,value=obj.Nota)	
		Libro.save(Ruta)

	else:
		Ruta = "/var/www/vhosts/colegio_venv/colegio/static/files/PLANTILLA_LIBRETA_PRIMARIA.xlsx"
		#Ruta = "static/files/PLANTILLA_LIBRETA_PRIMARIA.xlsx"

		Libro =load_workbook(Ruta)			
		Hoja1 = Libro.active
		fila=6
		col=3
		Hoja1["C3"]=""#Limpiando PeriodoAcademico Grado y Seccion
		for filas in range(7,65):#LIMPIANDO_FILAS
			Hoja1["B"+str(filas)]=""
			Hoja1.cell(row=4,column=filas-3,value="")
			Hoja1.cell(row=5,column=filas-3,value="")
			Hoja1.cell(row=6,column=filas-3,value="")
			for columnas in range(2,65):#LIMPIANDO_COLUMNAS
				Hoja1.cell(row=filas,column=columnas,value="")
			
		Hoja1["C3"]=periodo+" - "+gradonivel+" - "+seccion
		
		#poniendo los cursos y las competencias en los encabezados del excel
		cursocambio=''
		compe=0
		for cur in cursos:
			col+=1
			Hoja1.cell(row=4,column=col,value=cur.Curso.Nombre)
			Hoja1.cell(row=5,column=col,value=cur.Competencias.nombre_competencia)
			if cursocambio!=cur.Curso.Nombre:
				cursocambio=cur.Curso.Nombre
				compe=1
			else:
				compe+=1
			Hoja1.cell(row=6,column=col,value='C'+str(compe))
		
		colnota=3
		for mat in matriculas:#recorre cada fila
			fila+=1
			colnota+=1
			Hoja1["B"+str(fila)] = mat.Alumno.DNI
			Hoja1["C"+str(fila)] = mat.Alumno.ApellidoPaterno+" "+mat.Alumno.ApellidoMaterno+" "+mat.Alumno.Nombres
			for obj in consolidado_notas:
				if mat.id == obj.Matricula.id:
					for colcur in range(4,col+1):
						if obj.Curso.Nombre==Hoja1.cell(row=4,column=colcur).value and obj.Competencias.nombre_competencia==Hoja1.cell(row=5,column=colcur).value:##para saber si cambia curso
							Hoja1.cell(row=fila,column=colcur,value=obj.Nota)	
		Libro.save(Ruta)


def InsertaNotasEnExcelPromedio(consolidado_notas,matriculas,nivel,gradonivel,seccion,periodo,cursos):
	#LIMPIANDO TODAS LAS CELDAS################
	if str(nivel)=='SEC':		
		Ruta = "/var/www/vhosts/colegio_venv/colegio/static/files/PLANTILLA_LIBRETA_SECUNDARIA_R.xlsx"
		#Ruta = "static/files/PLANTILLA_LIBRETA_SECUNDARIA_R.xlsx"
		Libro =load_workbook(Ruta)			
		Hoja1 = Libro.active
		fila=6
		col=4
		Hoja1["C3"]=""#Limpiando PeriodoAcademico Grado y Seccion
		# for colu in range(4,11):
		# 	Hoja1.unmerge_cells(start_row=4,start_column=colu,end_row=4,end_column=colu+1)
		# 	Hoja1.unmerge_cells(start_row=5,start_column=colu,end_row=5,end_column=colu+1)

		for filas in range(7,65):#LIMPIANDO_FILAS
			Hoja1["B"+str(filas)]=""
			Hoja1["C"+str(filas)]=""
			Hoja1["D"+str(filas)]=""
			Hoja1["F"+str(filas)]=""
			Hoja1["H"+str(filas)]=""
			Hoja1["J"+str(filas)]=""
			Hoja1["L"+str(filas)]=""
			Hoja1["N"+str(filas)]=""
			Hoja1["P"+str(filas)]=""
			Hoja1["R"+str(filas)]=""
			Hoja1["T"+str(filas)]=""
			Hoja1["V"+str(filas)]=""
			Hoja1["X"+str(filas)]=""
			Hoja1["AB"+str(filas)]=""
			Hoja1["AD"+str(filas)]=""
		# for columnas in range(2,20):
		# 	Hoja1.cell(row=4,column=columnas+2,value="")
		# 	Hoja1.cell(row=5,column=columnas+2,value="")
		# 	Hoja1.cell(row=6,column=columnas+2,value="")
			
			# for columnas in range(2,65):#LIMPIANDO_COLUMNAS
			# Hoja1.cell(row=filas,column=columnas,value="")
			
		Hoja1["C3"]=periodo+" - "+gradonivel+" - "+seccion
		
		#poniendo los cursos y las competencias en los encabezados del excel
		cursocambio=''
		compe=0
		for cur in cursos:
			if cur.Competencias.nombre_competencia=='CALIFICATIVO DE ÁREA' :
				Hoja1.cell(row=4,column=col,value=cur.Curso.Nombre)
				Hoja1.cell(row=5,column=col,value=cur.Competencias.nombre_competencia)
				
				if cursocambio!=cur.Curso.Nombre:
					cursocambio=cur.Curso.Nombre
					compe=1
				else:
					compe+=1
				Hoja1.cell(row=6,column=col,value='C'+str(compe))
				col+=2

			if cur.Competencias.nombre_competencia=='APRECIACIÓN DEL TUTOR':
				Hoja1["AD4"]=cur.Curso.Nombre
				col+=2
				# Hoja1["AD5"]=cur.Competencias.nombre_competencia
			
			if cur.Competencias.nombre_competencia=='COMPORTAMIENTO':
				Hoja1["AB4"]=cur.Curso.Nombre
				col+=2
				# Hoja1["AB5"]=cur.Competencias.nombre_competencia
						
				# Hoja1.merge_cells(start_row=4,start_column=col,end_row=4,end_column=col+1)
				# Hoja1.merge_cells(start_row=5,start_column=col,end_row=5,end_column=col+1)
				
		
		colnota=3
		for mat in matriculas:#recorre cada fila
			fila+=1
			colnota+=1
			Hoja1["B"+str(fila)] = mat.Alumno.DNI
			Hoja1["C"+str(fila)] = mat.Alumno.ApellidoPaterno+" "+mat.Alumno.ApellidoMaterno+" "+mat.Alumno.Nombres
			for obj in consolidado_notas:
				if mat.id == obj.Matricula.id:
					for colcur in range(2,col+2):
						# print(Hoja1.cell(row=4,column=colcur).value)
						# print(Hoja1.cell(row=5,column=colcur).value)
						if obj.Curso.Nombre==Hoja1.cell(row=4,column=colcur).value and obj.Competencias.nombre_competencia==Hoja1.cell(row=5,column=colcur).value:##para saber si cambia curso
							Hoja1.cell(row=fila,column=colcur,value=obj.Nota)
		Libro.save(Ruta)

	else:
		Ruta = "/var/www/vhosts/colegio_venv/colegio/static/files/PLANTILLA_LIBRETA_PRIMARIA_R.xlsx"
		#Ruta = "static/files/PLANTILLA_LIBRETA_PRIMARIA_R.xlsx"
		Libro =load_workbook(Ruta)			
		Hoja1 = Libro.active
		fila=6
		col=4
		Hoja1["C3"]=""#Limpiando PeriodoAcademico Grado y Seccion
		for filas in range(7,55):#LIMPIANDO_FILAS
			Hoja1["B"+str(filas)]=""
			Hoja1["C"+str(filas)]=""
			Hoja1["D"+str(filas)]=""
			Hoja1["E"+str(filas)]=""
			Hoja1["F"+str(filas)]=""
			Hoja1["G"+str(filas)]=""
			Hoja1["H"+str(filas)]=""
			Hoja1["I"+str(filas)]=""
			Hoja1["J"+str(filas)]=""
			Hoja1["K"+str(filas)]=""
			Hoja1["L"+str(filas)]=""
			Hoja1["M"+str(filas)]=""
			Hoja1["N"+str(filas)]=""
			Hoja1["O"+str(filas)]=""
			Hoja1["P"+str(filas)]=""
			Hoja1["Q"+str(filas)]=""
			Hoja1["R"+str(filas)]=""
			Hoja1["S"+str(filas)]=""
			Hoja1["T"+str(filas)]=""
			Hoja1["U"+str(filas)]=""
			Hoja1["V"+str(filas)]=""
			Hoja1["W"+str(filas)]=""
			Hoja1["X"+str(filas)]=""
			Hoja1["Y"+str(filas)]=""
			Hoja1["Z"+str(filas)]=""
			Hoja1["AA"+str(filas)]=""
			
		Hoja1["C3"]=periodo+" - "+gradonivel+" - "+seccion
		
		#poniendo los cursos y las competencias en los encabezados del excel
		cursocambio=''
		compe=0
		for cur in cursos:
			if cur.Competencias.nombre_competencia=='CALIFICATIVO DE ÁREA' or cur.Curso.Tipo=='INASISTENCIAS' or cur.Curso.Tipo=='ACTITUDINAL' or cur.Curso.Tipo=='DEL PADRE DE FAMILIA'  or cur.Curso.Tipo=='APRECIACIÓN DEL TUTOR':
				if cur.Curso.Tipo=='CURSO':
					Hoja1.cell(row=4,column=col,value=cur.Curso.Nombre)
				else:
					Hoja1.cell(row=4,column=col,value=cur.Curso.Tipo)
				Hoja1.cell(row=5,column=col,value=cur.Competencias.nombre_competencia)
				col+=1
			if cursocambio!=cur.Curso.Nombre:
				cursocambio=cur.Curso.Nombre
				compe=1
			else:
				compe+=1
			Hoja1.cell(row=6,column=col,value='C'+str(compe))
		
		colnota=3
		for mat in matriculas:#recorre cada fila
			fila+=1
			colnota+=1
			Hoja1["B"+str(fila)] = mat.Alumno.DNI
			Hoja1["C"+str(fila)] = mat.Alumno.ApellidoPaterno+" "+mat.Alumno.ApellidoMaterno+" "+mat.Alumno.Nombres
			for obj in consolidado_notas:
				if mat.id == obj.Matricula.id:
					for colcur in range(4,col+1):
						if obj.Curso.Tipo=='CURSO':
							if obj.Curso.Nombre==Hoja1.cell(row=4,column=colcur).value and obj.Competencias.nombre_competencia==Hoja1.cell(row=5,column=colcur).value:##para saber si cambia curso
								Hoja1.cell(row=fila,column=colcur,value=obj.Nota)	
						else:
							if obj.Curso.Tipo==Hoja1.cell(row=4,column=colcur).value and obj.Competencias.nombre_competencia==Hoja1.cell(row=5,column=colcur).value:##para saber si cambia curso
								Hoja1.cell(row=fila,column=colcur,value=obj.Nota)	
		Libro.save(Ruta)
