from ast import For, Return
from distutils.command.sdist import sdist
from urllib import request, response

from django.shortcuts import render, redirect

from django.views.generic import ListView, DeleteView

from datetime import *

from django.db import connection ##permite conectar directamente a la base de datos

from openpyxl import Workbook,load_workbook

from colegio.Apps.AvanceNotas.forms import AvanceNotasForm,AvanceNotasCompForm

from colegio.Apps.Alumno.models import Alumno
from colegio.Apps.Curso.models import Curso
from colegio.Apps.AvanceNotas.models import AvanceNotas
from colegio.Apps.Matricula.models import Matricula
from colegio.Apps.AnoAcademico.models import AnoAcademico
from colegio.Apps.Docente.models import Docente
from colegio.Apps.DocenteCurso.models import DocenteCurso
from colegio.Apps.PeriodoAcademico.models import PAcademico
from colegio.Apps.AvanceTempDatos.models import AvanceTempDatos
from colegio.Apps.AvanceTempDatos.models import AvanceTempDatosComp
from colegio.Apps.Notas.models import SettingNotas
from colegio.Apps.Competencias.models import CompetenciaCurso, Competencias
from colegio.Apps.AvanceNotas.models import AvanceNotasComp
from django.contrib.auth.models import User
from django.http import JsonResponse

def ActualizarAvanceNotasPorAlumno(request):
	if request.method=='POST':
		matricula=request.POST.get('Matriculae')
		periodo=request.POST.get('PAcademicoe')
		curso=request.POST.get('Cursose')
		data=AvanceNotasComp.objects.filter(Matricula_id=matricula,PAcademico_id=periodo,Curso_id=curso)
		for dd in data:
			# print(request.POST.get("nota"+str(dd.id)))
			obj = AvanceNotasComp.objects.get(id=request.POST.get(str(dd.id)))
			obj.Nota = request.POST.get("nota"+str(dd.id))
			obj.save()

		data={'ok':'ok'}
		return JsonResponse(data)

def  ObtenerAvanceExistentesEditar(request):
	if request.method=='POST':
		matricula=request.POST.get('Matriculae')
		periodo=request.POST.get('PAcademicoe')
		data=list(AvanceNotasComp.objects.filter(Matricula_id=matricula,PAcademico_id=periodo).values())
		return JsonResponse(data,safe=False)

def  ObtenerAvanceExistentes(request):
	if request.method=='POST':
		matricula=request.POST.get('Matricula')
		periodo=request.POST.get('PAcademico')
		data=list(AvanceNotasComp.objects.filter(Matricula_id=matricula,PAcademico_id=periodo).values())
		return JsonResponse(data,safe=False)

def GuardaAvanceNotasPorAlumno(request):
	if request.method=='POST':
		idCurso=request.POST.get('Cursos')
		idDocente=Docente.objects.get(User__id=request.user.id)#Toma de Consola
		idMatricula=request.POST.get('Matricula')
		idPacademico=request.POST.get('PAcademico')		
		compes=Competencias.objects.filter(competenciacurso__Curso_id=idCurso).exclude(nombre_competencia='CALIFICATIVO DE ÁREA').order_by('Orden')
		for cc in compes:
			obj = AvanceNotasComp()
			obj.Competencias_id=cc.id
			obj.Curso_id=idCurso
			obj.Docente_id=idDocente.id
			obj.Matricula_id=idMatricula
			obj.PAcademico_id=idPacademico
			obj.Nota=request.POST.get(str(cc.id))
			obj.save()
		data={'ok':'ok'}
		return JsonResponse(data)
		
def ObtenerCompetenciasEditar(request):
	if request.method=='POST':
		matricula=request.POST.get('Matriculae')
		periodo=request.POST.get('PAcademicoe')
		curso=request.POST.get('Cursose')
		data=list(AvanceNotasComp.objects.filter(Matricula_id=matricula,PAcademico_id=periodo,Curso_id=curso).order_by('Competencias__Orden').values('id','Nota','Competencias__nombre_competencia'))
		return JsonResponse(data,safe=False)

def ObtenerCompetencias(request,idCurso):
	data=list(Competencias.objects.filter(competenciacurso__Curso_id=idCurso).exclude(nombre_competencia='CALIFICATIVO DE ÁREA').order_by('Orden').values())
	return JsonResponse(data,safe=False)

def BuscarCursoNivel(request, nivel):
	cursos_nivel=list(Curso.objects.filter(Tipo='CURSO',Nivel=nivel).order_by('Orden').values())#CURSOS
	return JsonResponse(cursos_nivel,safe=False)


def RegistroPorAlumno(request):
	#NIVEL SE VA DIRECTO EN EL FRONT
	aac=AnoAcademico.objects.all().order_by('-id')#AÑO
	pac=PAcademico.objects.filter(Status='Activo').order_by('id')#PERIODO ACADÉMICO
								#SECCIÓN ESTÁ EN EL FRONT
	cur=Curso.objects.filter(Tipo='CURSO',Nivel='PRIM').order_by('Orden')#CURSOS
								#GRADO NIVEL ESTÁ EN EL FRONT
	ano_actual=datetime.now().year
	alu=Matricula.objects.filter(AnoAcademico__Ano=ano_actual)
	contexto={'aac':aac,'pac':pac,'cur':cur,'alu':alu}
	return render(request,'avancenotas/RegistroNotasPorAlumno.html',contexto)

def ConsolidadoAvances(request):
	aac=AnoAcademico.objects.all().order_by('-id')
	pac=PAcademico.objects.all().order_by('id')
		
	if	request.method=='POST':
		ano=request.POST.get('ano')
		paca=request.POST.get('pacademico')#Determina la Vista que se seleccionará
		gradonivel=request.POST.get('Grado')
		seccion=request.POST.get('Seccion')
		
		nivel=str(gradonivel)[1:len(gradonivel)]
		for p in pac:
			if str(p.id)==paca:
				periodo=p.Nombre
		######con matriculas se obtiene todos los alumnos y se guarda en mtriculas de ese grado filtrado por ano,gradonivel, seccion
		####tambien se obtiene consolidado_notas de NotasComp
		matriculas=Matricula.objects.filter(AnoAcademico__Ano=ano,Grado=gradonivel,Seccion=seccion).order_by('Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres')
		consolidado_notas=AvanceNotasComp.objects.filter(Matricula__AnoAcademico__Ano=ano,
		Matricula__Grado=gradonivel,Matricula__Seccion=seccion,PAcademico=paca)
		cursos= CompetenciaCurso.objects.filter(Curso__Nivel=nivel,Curso__Tipo='CURSO').exclude(Competencias__nombre_competencia='CALIFICATIVO DE ÁREA').order_by('Curso__Orden','Competencias__Orden')
		###Mi funcion estrella que hace la granputa colocacion en el excel
		InsertaNotasEnExcel(consolidado_notas,matriculas,nivel,gradonivel,seccion,periodo,cursos)#funcion que coloca las notas en el excel se encuentra mas abajo
		############################################3
		if str(nivel).find("PRIM")!= -1:
			return redirect("https://colcoopcv.com/static/files/PLANTILLA_AVANCE_PRIMARIA.xlsx")
			#return redirect("/static/files/PLANTILLA_AVANCE_PRIMARIA.xlsx")
		else:
			return redirect("https://colcoopcv.com/static/files/PLANTILLA_AVANCE_SECUNDARIA.xlsx")
			#return redirect("/static/files/PLANTILLA_AVANCE_SECUNDARIA.xlsx")
	else:
		contexto={'aac':aac,'pac':pac}
		return render(request,'otras_opciones/imprimir_consolidado_avances.html',contexto)

def AvanceNotasNuevoComp(request):
	persona=AvanceTempDatosComp.objects.filter(User=request.user.id)
	if persona.exists():
		AvanceTempDatosComp.objects.filter(User=request.user.id).delete()

	ano = AnoAcademico.objects.get(Ano=datetime.now().year)#Año
	paca = PAcademico.objects.get(FechaInicio__lt=datetime.now(),FechaFinal__gt=datetime.now())#Bimestre
	doce = Docente.objects.get(User__id=request.user.id)#Profesor

	curso_list = DocenteCurso.objects.filter(Docente__User__id=request.user.id)#para refencia en GradoNivel plantilla Seccion
	
	docente = Docente.objects.get(User__id=request.user.id)
	grados_list = str(docente.GradoNivel).split()#Extrae los grados de la cadena y los pone en lista
	secciones_list = str(docente.Seccion).split()#Extrae las secciones de la cadena y los pone en lista
	######################
	#este contexto es para el ELSE
	contexto2 = {'doce':doce,'paca':paca,'ano':ano,'curso_list':curso_list,'grados_list':grados_list,'secciones_list':secciones_list} #'doce':doce}#'matri':matri
	if request.method=='POST':
		temp_datos = AvanceTempDatosComp()
		
		usu=User()
		usu.id=request.user.id
		doce.User=usu
		
		temp_datos.User = usu#nuevo
		temp_datos.idCurso = request.POST.get("curso")
		temp_datos.grado = request.POST.get("grados")
		temp_datos.seccion = request.POST.get("secciones")
		temp_datos.save()

		return redirect('app_avancenota_nuevo_save_comp')#aqui guarda los avances de notas primaria
	else:
		return render(request,'avancenotas/create_avancenotas_comp.html',contexto2)

def AvanceNotasNuevoSaveComp(request):	
	ano = AnoAcademico.objects.get(Ano=datetime.now().year)#Año
	paca = PAcademico.objects.get(FechaInicio__lt=datetime.now(),FechaFinal__gt=datetime.now())#Bimestre
	doce = Docente.objects.get(User__id=request.user.id)#Profesor
	
	usu=User()
	usu.id=request.user.id
	doce.User=usu

	#td = AvanceTempDatos.objects.get(User=usu) #funcionaba
	td_comp=AvanceTempDatosComp.objects.get(User=usu)
	
	#compe = Competencias.objects.get(id=td_comp.idCurso)
	curso = Curso.objects.get(id=td_comp.idCurso)

	grado =  td_comp.grado
	seccion = td_comp.seccion	
	notas = AvanceNotasComp.objects.filter(Curso=td_comp.idCurso, Matricula__Seccion=td_comp.seccion, Matricula__Grado=td_comp.grado, Matricula__AnoAcademico=ano.id, PAcademico__id=paca.id)
	#notas = AvanceNotas.objects.filter(Curso__id=td.idCurso, Matricula__Seccion=td.seccion, Matricula__Grado=td.grado, Matricula__AnoAcademico=ano.id, PAcademico__id=paca.id)
	#Encontrando las competencias del Curso
	lista_compe=CompetenciaCurso.objects.filter(Curso__id=curso.id).exclude(Competencias__nombre_competencia='CALIFICATIVO DE ÁREA').order_by('Competencias__Orden')
	######################################################################

	if notas:
		alumnos = ''		
	else:
		alumnos = Matricula.objects.filter(Grado=td_comp.grado,Seccion=td_comp.seccion,AnoAcademico=ano.id).order_by('Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres')
	
	contexto={'grado':grado,'seccion':seccion,'curso':curso,'ano':ano,'paca':paca,'doce':doce,'alumnos':alumnos,'lista_compe':lista_compe}
	
	if request.method=='POST':
		for alu in alumnos:
			for cc in lista_compe:
				notacomp = AvanceNotasComp()

				cur = Curso()
				cur.id = td_comp.idCurso
				notacomp.Curso = cur

				comp = Competencias()
				comp.id = cc.Competencias.id
				notacomp.Competencias = comp

				pac = PAcademico()
				pac.id = paca.id
				notacomp.PAcademico = pac

				doc = Docente()
				doc.id = doce.id
				notacomp.Docente = doc
		
				mat = Matricula()
				mat.id = alu.id
				notacomp.Matricula = mat

				#es alu.id porque el inputtextNota lleva el nombre del id de matricula
				evalu=alu.id #obtener el id del alumno
				evcomp=cc.Competencias.id##obtener el id de la competencia para combinarlo
				
				nota_eva = request.POST.get(str(evalu)+str(evcomp))
				if nota_eva=='':
					nota_eva='COMPETENCIA NO DESARROLLADA'
				notacomp.Nota = str(nota_eva).upper()
				notacomp.save()
		grabo='registrado'
		
		notita = AvanceNotasComp.objects.filter(Docente__User__id=request.user.id, Curso__id=td_comp.idCurso, Matricula__Grado=td_comp.grado, Matricula__Seccion=td_comp.seccion,Matricula__AnoAcademico=ano.id,PAcademico__id=paca.id).order_by('Matricula__Alumno__ApellidoPaterno','Matricula__Alumno__ApellidoMaterno','Matricula__Alumno__Nombres')
		context = {'grabo':grabo,'curso':curso,'grado':grado,'seccion':seccion,'notita':notita}

		persona=AvanceTempDatosComp.objects.filter(User=request.user.id)
		if persona.exists():
			AvanceTempDatos.objects.filter(User=request.user.id).delete()
		return render(request,'avancenotas/listar_avancenotas.html', context)
	else:
		return render(request,'avancenotas/create_avancenotas_save_comp.html',contexto)




def AvanceNotasNuevo(request):
	#primero debe borrar los datos de la tabla del usuario actual
	#TempDatos.objects.filter(idusuario=request.user.id).delete()
	persona=AvanceTempDatos.objects.filter(User=request.user.id)
	if persona.exists():
		AvanceTempDatos.objects.filter(User=request.user.id).delete()

	ano = AnoAcademico.objects.get(Ano=datetime.now().year)#Año
	paca = PAcademico.objects.get(FechaInicio__lt=datetime.now(),FechaFinal__gt=datetime.now())#Bimestre
	doce = Docente.objects.get(User__id=request.user.id)#Profesor

	curso_list = DocenteCurso.objects.filter(Docente__User__id=request.user.id)#para refencia en GradoNivel plantilla Seccion

	docente = Docente.objects.get(User__id=request.user.id)
	grados_list = str(docente.GradoNivel).split()#Extrae los grados de la cadena y los pone en lista
	secciones_list = str(docente.Seccion).split()#Extrae las secciones de la cadena y los pone en lista
	######################
	#este contexto es para el ELSE
	contexto2 = {'doce':doce,'paca':paca,'ano':ano,'curso_list':curso_list,'grados_list':grados_list,'secciones_list':secciones_list} #'doce':doce}#'matri':matri
	if request.method=='POST':
		temp_datos = AvanceTempDatos()
		
		usu=User()
		usu.id=request.user.id
		doce.User=usu
		
		temp_datos.User = usu#nuevo
		temp_datos.idCurso = request.POST.get("curso")
		temp_datos.grado = request.POST.get("grados")
		temp_datos.seccion = request.POST.get("secciones")
		temp_datos.save()

		return redirect('app_avancenota_nuevo_save')
	else:
		return render(request,'avancenotas/create_avancenotas.html',contexto2)

def AvanceNotasNuevoSave(request):	
	ano = AnoAcademico.objects.get(Ano=datetime.now().year)#Año
	paca = PAcademico.objects.get(FechaInicio__lt=datetime.now(),FechaFinal__gt=datetime.now())#Bimestre
	doce = Docente.objects.get(User__id=request.user.id)#Profesor
	
	#tempd = TempDatos.objects.latest('id') #funcionaba
	usu=User()
	usu.id=request.user.id
	doce.User=usu

	td = AvanceTempDatos.objects.get(User=usu) #funcionaba
	curso = Curso.objects.get(id=td.idCurso)
	grado =  td.grado
	seccion = td.seccion	
	notas = AvanceNotas.objects.filter(Curso__id=td.idCurso, Matricula__Seccion=td.seccion, Matricula__Grado=td.grado, Matricula__AnoAcademico=ano.id, PAcademico__id=paca.id)

	if notas:
		alumnos = ''		
	else:
		alumnos = Matricula.objects.filter(Grado=td.grado,Seccion=td.seccion,AnoAcademico=ano.id).order_by('Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres')
	
	contexto={'grado':grado,'seccion':seccion,'curso':curso,'ano':ano,'paca':paca,'doce':doce,'alumnos':alumnos}
	
	if request.method=='POST':

		for alu in alumnos:
			nota = AvanceNotas()
			cur = Curso()
			cur.id = td.idCurso
			nota.Curso = cur

			pac = PAcademico()
			pac.id = paca.id
			nota.PAcademico = pac

			doc = Docente()
			doc.id = doce.id
			nota.Docente = doc
	
			mat = Matricula()
			mat.id = alu.id
			nota.Matricula = mat

			#es alu.id porque el inputtextNota lleva el nombre del id de matricula
			evalu=alu.id
			nota_eva = request.POST.get(str(evalu))
			simulacronota_eva=request.POST.get(str("simu"+str(evalu)))
			nota.Nota = str(nota_eva).upper()
			nota.SimulacroNota = str(simulacronota_eva).upper()
			nota.save()
		grabo='registrado'
		
		notita = AvanceNotas.objects.filter(Docente__User__id=request.user.id, Curso__id=td.idCurso, Matricula__Grado=td.grado, Matricula__Seccion=td.seccion,Matricula__AnoAcademico=ano.id,PAcademico__id=paca.id).order_by('Matricula__Alumno__ApellidoPaterno','Matricula__Alumno__ApellidoMaterno','Matricula__Alumno__Nombres')
		context = {'grabo':grabo,'curso':curso,'grado':grado,'seccion':seccion,'notita':notita}

		persona=AvanceTempDatos.objects.filter(User=request.user.id)
		if persona.exists():
			AvanceTempDatos.objects.filter(User=request.user.id).delete()
		return render(request,'avancenotas/listar_avancenotas.html', context)
	else:
		return render(request,'avancenotas/create_avancenotas_save.html',contexto)

def AvanceNotasNuevoUno(request):
	ano = AnoAcademico.objects.get(Ano=datetime.now().year)

	paca = PAcademico.objects.get(FechaInicio__lt=datetime.now(),FechaFinal__gt=datetime.now())#Bimestre
	doce = Docente.objects.get(User__id=request.user.id)#Profesor
	curso = Curso.objects.all()
	matri = Matricula.objects.filter(AnoAcademico=ano.id)
	contexto={'curso':curso,'matri':matri}
	
	if request.method=='POST':
		avancenotas = AvanceNotasComp()

		pac = PAcademico()
		pac.id = paca.id
		avancenotas.PAcademico = pac

		doc = Docente()
		doc.id = doce.id
		avancenotas.Docente = doc

		cur = Curso()
		cur.id = request.POST.get("Curso")
		avancenotas.Curso = cur
		
		mat = Matricula()
		mat.id = request.POST.get("Alumno")
		avancenotas.Matricula = mat

		avancenotas.Nota = request.POST.get("Nota")
		avancenotas.SimulacroNota = request.POST.get("SimulacroNota")
		avancenotas.save()
		return redirect('app_listar_avancenotas')
	else:
		return render(request,'avancenotas/create_avancenotas_save_uno.html',contexto)


def AvanceNotasEdit(request, id_notas):
	nota = AvanceNotasComp.objects.get(id=id_notas)
	if request.method =='GET':
		form = AvanceNotasCompForm(instance=nota)
	else:
		form = AvanceNotasCompForm(request.POST,instance=nota)
		if form.is_valid():
			form.save()
		# return redirect('app_listar_avancenotas')#tambien se puede usar return redirect('app_listar_notas')
	contexto = {'nota':nota,'form':form}
	return render(request, 'avancenotas/editar_avancenotas.html',contexto)

class AvanceNotasDelete(DeleteView):
	model=AvanceNotasComp
	template_name = 'avancenotas/delete_avancenotas.html'
	success_url ='/avancenotas/listar/' 


def EstadoValor(obj):
	if obj:#Si es True
		devuelve='checked=""'
	else:
		devuelve=''

	return (devuelve)

def AvanceListaNotas(request):
	ingnotas=SettingNotas.objects.get(id=1)
	ingavances=SettingNotas.objects.get(id=2)
	ingnotas=EstadoValor(ingnotas.Valor)
	ingavances=EstadoValor(ingavances.Valor)
	
	if ingnotas=='checked=""':
		chk_bimestre='on'
	else:
		chk_bimestre=''

	if ingavances=='checked=""':
		chk_avance='on'
	else:
		chk_avance=''
	
	contexto={'ingnotas':ingnotas, 'ingavances':ingavances,'chk_bimestre':chk_bimestre,'chk_avance':chk_avance}
	return render(request, 'avancenotas/listar_avancenotas.html',contexto)
def DeleteAvanceNotasxCurso(request):
	Oano=AnoAcademico.objects.all()
	Opac=PAcademico.objects.all()
	Ocur=Curso.objects.all()
	Omsje=''
	contexto={'ano':Oano,'pac':Opac,'cur':Ocur,'msje':Omsje}
	if request.method=='POST':
		#Eliminar las Notas x Curso que se envia a traves del POST
		cur=Curso()
		ano=AnoAcademico()
		pac=PAcademico()

		ano.id=request.POST.get("Ano")
		pac.id=request.POST.get("PAcademico")
		cur.id=request.POST.get("Curso")
		gra=request.POST.get("Grado")
		sec=request.POST.get("Seccion")

		AvanceNotasComp.objects.filter(Matricula__AnoAcademico=ano.id, PAcademico__id=pac.id,Curso=cur.id, Matricula__Grado=gra,Matricula__Seccion=sec).delete()
		
		Omsje='eliminado'
		contexto={'ano':Oano,'pac':Opac,'cur':Ocur,'msje':Omsje}
		return render(request,'avancenotas/delete_avancenotas_xcurso.html',contexto)
	else:
		msje='-'
		return render(request,'avancenotas/delete_avancenotas_xcurso.html',contexto)

	
def InsertaNotasEnExcel(consolidado_notas,matriculas,nivel,gradonivel,seccion,periodo,cursos):
	if str(nivel)=='SEC':		
		Ruta = "/var/www/vhosts/colegio_venv/colegio/static/files/PLANTILLA_AVANCE_SECUNDARIA.xlsx"
		#Ruta = "static/files/PLANTILLA_AVANCE_SECUNDARIA.xlsx"
		Libro =load_workbook(Ruta)			
		Hoja1 = Libro.active
		fila=6
		col=3
		Hoja1["C3"]=""#Limpiando PeriodoAcademico Grado y Seccion
		for filas in range(7,65):#LIMPIANDO_FILAS
			Hoja1["B"+str(filas)]=""
			Hoja1.cell(row=4,column=filas-3,value="")
			Hoja1.cell(row=5,column=filas-3,value="")
			Hoja1.cell(row=6,column=filas-3,value="")
			for columnas in range(2,65):#LIMPIANDO_COLUMNAS
				Hoja1.cell(row=filas,column=columnas,value="")
			
		Hoja1["C3"]=periodo+" - "+gradonivel+" - "+seccion
		
		#poniendo los cursos y las competencias en los encabezados del excel
		cursocambio=''
		compe=0
		for cur in cursos:
			col+=1
			Hoja1.cell(row=4,column=col,value=cur.Curso.Nombre)
			Hoja1.cell(row=5,column=col,value=cur.Competencias.nombre_competencia)
			if cursocambio!=cur.Curso.Nombre:
				cursocambio=cur.Curso.Nombre
				compe=1
			else:
				compe+=1
			Hoja1.cell(row=6,column=col,value='C'+str(compe))
		
		colnota=3
		for mat in matriculas:#recorre cada fila
			fila+=1
			colnota+=1
			Hoja1["B"+str(fila)] = mat.Alumno.DNI
			Hoja1["C"+str(fila)] = mat.Alumno.ApellidoPaterno+" "+mat.Alumno.ApellidoMaterno+" "+mat.Alumno.Nombres
			for obj in consolidado_notas:
				if mat.id == obj.Matricula.id:
					for colcur in range(4,col+1):
						if obj.Curso.Nombre==Hoja1.cell(row=4,column=colcur).value and obj.Competencias.nombre_competencia==Hoja1.cell(row=5,column=colcur).value:##para saber si cambia curso
							Hoja1.cell(row=fila,column=colcur,value=obj.Nota)	
		Libro.save(Ruta)
	else:
		#Ruta de Primaria
		Ruta = "/var/www/vhosts/colegio_venv/colegio/static/files/PLANTILLA_AVANCE_PRIMARIA.xlsx"
		#Ruta = "static/files/PLANTILLA_AVANCE_PRIMARIA.xlsx"
		Libro =load_workbook(Ruta)			
		Hoja1 = Libro.active
		fila=6
		col=3
		Hoja1["C3"]=""#Limpiando PeriodoAcademico Grado y Seccion
		for filas in range(7,65):#LIMPIANDO_FILAS
			Hoja1["B"+str(filas)]=""
			Hoja1.cell(row=4,column=filas-3,value="")
			Hoja1.cell(row=5,column=filas-3,value="")
			Hoja1.cell(row=6,column=filas-3,value="")
			for columnas in range(2,65):#LIMPIANDO_COLUMNAS
				Hoja1.cell(row=filas,column=columnas,value="")
			
		Hoja1["C3"]=periodo+" - "+gradonivel+" - "+seccion
		
		#poniendo los cursos y las competencias en los encabezados del excel
		cursocambio=''
		compe=0
		for cur in cursos:
			col+=1
			Hoja1.cell(row=4,column=col,value=cur.Curso.Nombre)
			Hoja1.cell(row=5,column=col,value=cur.Competencias.nombre_competencia)
			if cursocambio!=cur.Curso.Nombre:
				cursocambio=cur.Curso.Nombre
				compe=1
			else:
				compe+=1
			Hoja1.cell(row=6,column=col,value='C'+str(compe))
		
		colnota=3
		for mat in matriculas:#recorre cada fila
			fila+=1
			colnota+=1
			Hoja1["B"+str(fila)] = mat.Alumno.DNI
			Hoja1["C"+str(fila)] = mat.Alumno.ApellidoPaterno+" "+mat.Alumno.ApellidoMaterno+" "+mat.Alumno.Nombres
			for obj in consolidado_notas:
				if mat.id == obj.Matricula.id:
					for colcur in range(4,col+1):
						if obj.Curso.Nombre==Hoja1.cell(row=4,column=colcur).value and obj.Competencias.nombre_competencia==Hoja1.cell(row=5,column=colcur).value:##para saber si cambia curso
							Hoja1.cell(row=fila,column=colcur,value=obj.Nota)	
		Libro.save(Ruta)
