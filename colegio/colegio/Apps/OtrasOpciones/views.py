from ast import For
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse

from openpyxl import Workbook,load_workbook

from colegio.Apps.Matricula.models import Matricula
from colegio.Apps.Alumno.models import Alumno
from colegio.Apps.Curso.models import Curso
from colegio.Apps.Notas.models import Notas
from colegio.Apps.AvanceNotas.models import AvanceNotas,AvanceNotasComp
from colegio.Apps.Competencias.models import CompetenciaCurso
from colegio.Apps.AnoAcademico.models import AnoAcademico
from colegio.Apps.Docente.models import Docente
from colegio.Apps.PeriodoAcademico.models import PAcademico

from colegio.Apps.Matricula.forms import MatriculaForm
from colegio.Apps.Notas.forms import NotasForm
from colegio.Apps.Alumno.forms import AlumnoForm
from colegio.Apps.Docente.forms import DocenteForm
from colegio.Apps.Curso.forms import CursoForm
from colegio.Apps.PeriodoAcademico.forms import PAcademicoForm
from colegio.Apps.AnoAcademico.forms import AnoAcademicoForm

from math import *
from io import BytesIO

# from reportlab.pdfgen import canvas
# from reportlab.lib.pagesizes import letter,A4
# from reportlab.lib.pagesizes import landscape,portrait

# #Esto hace que lea imágenes
# from reportlab.lib.utils import ImageReader#Esto fue agregad

# from reportlab.platypus import Image
# from reportlab.lib.colors import HexColor #para cambiar colo texto en hexadecimales

from datetime import *
from django.contrib.auth.decorators import login_required

# def OpcionImprimir(request):
# 	return render(request,'otras_opciones/opcion_imprimir.html')

# def ConsolidadoAvances(request):
# 	return render(request,'otras_opciones/imprimir_consolidado_avances.html')

# def ConsolidadoLibretas(request):
# 	return render(request,'otras_opciones/imprimir_consolidado_libretas.html')

def ResumenAnual(request):
	return render(request,'otras_opciones/imprimir_resumen_anual.html')

def ImprimirResumenAnual(request):
	aac = AnoAcademico.objects.all().order_by('-Ano')
	pac = PAcademico.objects.all().order_by('Nombre')
	contexto = {'aac':aac,'pac':pac}
	if request.method == 'POST':
		grado=request.POST.get("Grado")
		seccion=request.POST.get("Seccion")

		ano=AnoAcademico()
		ano.id=request.POST.get("Ano")
		
		####################################################################
		matri = Matricula.objects.filter(Grado=grado,Seccion=seccion,AnoAcademico=ano).order_by('Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres')
		notas = Notas.objects.filter(Matricula__Grado=grado,Matricula__Seccion=seccion,Matricula__AnoAcademico=ano)
		if str(grado).find("PRIM")!= -1:			
			#Ruta de Primaria
			Ruta = "/static/files/RESUMEN_FINAL_PRIMARIA.xlsx"
		else:
			#Ruta de Secundaria 1SEC o 2SEC
			if str(grado).find("1SEC")!= -1 or str(grado).find("2SEC")!= -1 or str(grado).find("3SEC")!= -1:
				Ruta = "/static/files/RESUMEN_FINAL_SECUNDARIA_1_2.xlsx"
			else:
				#Ruta de 4SEC,5SEC
				Ruta = "/static/files/RESUMEN_FINAL_SECUNDARIA.xlsx"
		Libro = load_workbook(Ruta)
		Hoja1 = Libro.active
		Hoja1["B2"]=""
		if str(grado).find("PRIM")!=-1:
			for filas in range(4,50):#LIMPIANDO_FILAS
				Hoja1["B"+str(filas)]=""
				for columnas in range(2,50):#LIMPIANDO_COLUMNAS
					Hoja1.cell(row=filas,column=columnas,value="")

		if str(grado).find("SEC")!=-1:
			for filas in range(4,50):#LIMPIANDO_FILAS
				Hoja1["B"+str(filas)]=""
				
				Hoja1["C"+str(filas)]=""
				Hoja1["E"+str(filas)]=""
				Hoja1["G"+str(filas)]=""
				Hoja1["I"+str(filas)]=""
				
				Hoja1["L"+str(filas)]=""
				Hoja1["N"+str(filas)]=""
				Hoja1["P"+str(filas)]=""
				Hoja1["R"+str(filas)]=""
				
				Hoja1["U"+str(filas)]=""
				Hoja1["W"+str(filas)]=""
				Hoja1["Y"+str(filas)]=""
				Hoja1["AA"+str(filas)]=""

				Hoja1["AD"+str(filas)]=""
				Hoja1["AF"+str(filas)]=""
				Hoja1["AH"+str(filas)]=""
				Hoja1["AJ"+str(filas)]=""

				Hoja1["AM"+str(filas)]=""
				Hoja1["AO"+str(filas)]=""
				Hoja1["AQ"+str(filas)]=""
				Hoja1["AS"+str(filas)]=""

				Hoja1["AV"+str(filas)]=""
				Hoja1["AX"+str(filas)]=""
				Hoja1["AZ"+str(filas)]=""
				Hoja1["BB"+str(filas)]=""

				Hoja1["BE"+str(filas)]=""
				Hoja1["BG"+str(filas)]=""
				Hoja1["BI"+str(filas)]=""
				Hoja1["BK"+str(filas)]=""

				Hoja1["BN"+str(filas)]=""
				Hoja1["BP"+str(filas)]=""
				Hoja1["BR"+str(filas)]=""
				Hoja1["BT"+str(filas)]=""

				Hoja1["BW"+str(filas)]=""
				Hoja1["BY"+str(filas)]=""
				Hoja1["CA"+str(filas)]=""
				Hoja1["CC"+str(filas)]=""

				Hoja1["CF"+str(filas)]=""
				Hoja1["CH"+str(filas)]=""
				Hoja1["CJ"+str(filas)]=""
				Hoja1["CL"+str(filas)]=""

				Hoja1["CO"+str(filas)]=""
				Hoja1["CQ"+str(filas)]=""
				Hoja1["CS"+str(filas)]=""
				Hoja1["CU"+str(filas)]=""
				#A partir de aqui es seguidito
				Hoja1["CY"+str(filas)]=""
				Hoja1["CZ"+str(filas)]=""
				Hoja1["DA"+str(filas)]=""
				Hoja1["DB"+str(filas)]=""
				Hoja1["DC"+str(filas)]=""
				Hoja1["DD"+str(filas)]=""
				Hoja1["DE"+str(filas)]=""
				Hoja1["DF"+str(filas)]=""
				Hoja1["DG"+str(filas)]=""
				Hoja1["DH"+str(filas)]=""
				Hoja1["DI"+str(filas)]=""
				Hoja1["DJ"+str(filas)]=""
				Hoja1["DK"+str(filas)]=""
				Hoja1["DL"+str(filas)]=""
				Hoja1["DM"+str(filas)]=""
				Hoja1["DN"+str(filas)]=""
				Hoja1["DO"+str(filas)]=""
				Hoja1["DP"+str(filas)]=""
				Hoja1["DQ"+str(filas)]=""
				Hoja1["DR"+str(filas)]=""
				Hoja1["DS"+str(filas)]=""
				Hoja1["DT"+str(filas)]=""
				Hoja1["DU"+str(filas)]=""
				Hoja1["DV"+str(filas)]=""
				Hoja1["DW"+str(filas)]=""
		else:
			for filas in range(4,50):#LIMPIANDO_FILAS
				Hoja1["B"+str(filas)]=""
				
				Hoja1["C"+str(filas)]=""
				Hoja1["D"+str(filas)]=""
				Hoja1["E"+str(filas)]=""
				Hoja1["F"+str(filas)]=""
				
				Hoja1["G"+str(filas)]=""
				Hoja1["H"+str(filas)]=""
				Hoja1["I"+str(filas)]=""
				Hoja1["J"+str(filas)]=""
				
				Hoja1["K"+str(filas)]=""
				Hoja1["L"+str(filas)]=""
				Hoja1["M"+str(filas)]=""
				Hoja1["N"+str(filas)]=""

				Hoja1["O"+str(filas)]=""
				Hoja1["P"+str(filas)]=""
				Hoja1["Q"+str(filas)]=""
				Hoja1["R"+str(filas)]=""

				Hoja1["S"+str(filas)]=""
				Hoja1["T"+str(filas)]=""
				Hoja1["U"+str(filas)]=""
				Hoja1["V"+str(filas)]=""

				Hoja1["W"+str(filas)]=""
				Hoja1["X"+str(filas)]=""
				Hoja1["Y"+str(filas)]=""
				Hoja1["Z"+str(filas)]=""

				Hoja1["AA"+str(filas)]=""
				Hoja1["AB"+str(filas)]=""
				Hoja1["AC"+str(filas)]=""
				Hoja1["AD"+str(filas)]=""

				Hoja1["AE"+str(filas)]=""
				Hoja1["AF"+str(filas)]=""
				Hoja1["AG"+str(filas)]=""
				Hoja1["AH"+str(filas)]=""

				Hoja1["AI"+str(filas)]=""
				Hoja1["AJ"+str(filas)]=""
				Hoja1["AK"+str(filas)]=""
				Hoja1["AL"+str(filas)]=""

				Hoja1["AM"+str(filas)]=""
				Hoja1["AN"+str(filas)]=""
				Hoja1["AO"+str(filas)]=""
				Hoja1["AP"+str(filas)]=""

				Hoja1["AQ"+str(filas)]=""
				Hoja1["AR"+str(filas)]=""
				Hoja1["AS"+str(filas)]=""
				Hoja1["AT"+str(filas)]=""
				#A partir de aqui es seguidito
				Hoja1["AU"+str(filas)]=""
				Hoja1["AV"+str(filas)]=""
				Hoja1["AW"+str(filas)]=""
				Hoja1["AX"+str(filas)]=""
				Hoja1["AY"+str(filas)]=""
				Hoja1["AZ"+str(filas)]=""
				
				Hoja1["BA"+str(filas)]=""
				Hoja1["BB"+str(filas)]=""
				Hoja1["BC"+str(filas)]=""
				Hoja1["BD"+str(filas)]=""
				Hoja1["BE"+str(filas)]=""
				Hoja1["BF"+str(filas)]=""
				Hoja1["BG"+str(filas)]=""
				Hoja1["BH"+str(filas)]=""
				Hoja1["BI"+str(filas)]=""
				Hoja1["BJ"+str(filas)]=""
				Hoja1["BK"+str(filas)]=""
				Hoja1["BL"+str(filas)]=""
				Hoja1["BM"+str(filas)]=""
				Hoja1["BN"+str(filas)]=""
				Hoja1["BO"+str(filas)]=""
				Hoja1["BP"+str(filas)]=""
				Hoja1["BQ"+str(filas)]=""
				Hoja1["BR"+str(filas)]=""
				Hoja1["BS"+str(filas)]=""

				Hoja1["BT"+str(filas)]=""
				Hoja1["BU"+str(filas)]=""
				Hoja1["BV"+str(filas)]=""
				Hoja1["BW"+str(filas)]=""
				Hoja1["BX"+str(filas)]=""
				Hoja1["BY"+str(filas)]=""
				Hoja1["BZ"+str(filas)]=""

				Hoja1["CA"+str(filas)]=""
				Hoja1["CB"+str(filas)]=""
				Hoja1["CC"+str(filas)]=""
				Hoja1["CD"+str(filas)]=""
				Hoja1["CE"+str(filas)]=""
				Hoja1["CF"+str(filas)]=""
				Hoja1["CG"+str(filas)]=""
				Hoja1["CH"+str(filas)]=""
				Hoja1["CI"+str(filas)]=""
				Hoja1["CJ"+str(filas)]=""
				Hoja1["CK"+str(filas)]=""
				Hoja1["CL"+str(filas)]=""
				Hoja1["CM"+str(filas)]=""
				Hoja1["CN"+str(filas)]=""
				Hoja1["CO"+str(filas)]=""
				Hoja1["CP"+str(filas)]=""
				Hoja1["CQ"+str(filas)]=""
				Hoja1["CR"+str(filas)]=""
				Hoja1["CS"+str(filas)]=""
				Hoja1["CT"+str(filas)]=""
				Hoja1["CU"+str(filas)]=""
				Hoja1["CV"+str(filas)]=""
				Hoja1["CW"+str(filas)]=""
				Hoja1["CX"+str(filas)]=""

		valor=3
		for m in matri:
			valor = valor + 1 #Empieza por la fila 4
			Hoja1["B"+str(valor)] = m.Alumno.ApellidoPaterno + " " + m.Alumno.ApellidoMaterno + " " + m.Alumno.Nombres
			for n in notas:

				Hoja1["B2"]= str(n.Matricula.AnoAcademico.Ano) + " - " + grado+" - "+seccion
				if (m.id==n.Matricula.id):
					if str(grado).find("SEC")!= -1:
						if str(n.Curso.Nombre).find("MATEMÁTICA")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["C"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["E"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["G"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["I"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("COMUNICACIÓN")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["L"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["N"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["P"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["R"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("INGLÉS")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["U"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["W"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["Y"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["AA"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("CHINO MANDARÍN")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["AD"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["AF"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["AH"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["AJ"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("ARTE Y CULTURA")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["AM"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["AO"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["AQ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["AS"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("CIENCIAS SOCIALES")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["AV"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["AX"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["AZ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["BB"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("DESARROLLO PERSONAL")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["BE"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["BG"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["BI"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["BK"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("EDUCACIÓN FÍSICA")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["BN"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["BP"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["BR"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["BT"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("EDUCACIÓN RELIGIOSA")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["BW"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["BY"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["CA"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["CC"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("CIENCIA Y TECNOLOGÍA")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["CF"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["CH"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["CJ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["CL"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("EDUCACIÓN PARA EL TRABAJO")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["CO"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["CQ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["CS"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["CU"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("COMPORTAMIENTO")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["CZ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["DA"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["DB"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["DC"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("ORDEN DE MÉRITO")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["DD"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["DE"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["DF"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["DG"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("JUSTIFICADAS")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["DH"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["DI"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["DJ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["DK"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("INJUSTIFICADAS")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["DL"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["DM"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["DN"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["DO"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("TARDANZAS")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["DP"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["DQ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["DR"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["DS"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("APRECIACI")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["DT"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["DU"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["DV"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["DW"+str(valor)] = n.Nota
					else:#PRIMARIA
						
						if str(n.Curso.Nombre).find("MATEMÁTICA")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["C"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["D"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["E"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["F"+str(valor)] = n.Nota

						if str(n.Curso.Nombre).find("COMUNICACIÓN")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["G"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["H"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["I"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["J"+str(valor)] = n.Nota

						if str(n.Curso.Nombre).find("ARTE Y CULTURA")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["K"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["L"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["M"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["N"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("PERSONAL SOCIAL")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["O"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["P"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["Q"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["R"+str(valor)] = n.Nota

						if str(n.Curso.Nombre).find("EDUCACIÓN FÍSICA")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["S"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["T"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["U"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["V"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("EDUCACIÓN RELIGIOSA")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["W"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["X"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["Y"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["Z"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("CIENCIA Y TECNOLOGÍA")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["AA"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["AB"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["AC"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["AD"+str(valor)] = n.Nota

						if str(n.Curso.Nombre).find("INGLÉS")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["AE"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["AF"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["AG"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["AH"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("CHINO MANDARÍN")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["AI"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["AJ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["AK"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["AL"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("COMPUTACIÓN")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["AM"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["AN"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["AO"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["AP"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("RESPETO")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["AQ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["AR"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["AS"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["AT"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("PUNTUALIDAD")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["AU"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["AV"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["AW"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["AX"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("ORDEN")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["AY"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["AZ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["BA"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["BB"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("DISCIPLINA")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["BC"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["BD"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["BE"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["BF"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("PARTICIPACIÓN")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["BG"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["BH"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["BI"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["BJ"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("JUSTIFICADAS")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["BK"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["BL"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["BM"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["BN"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("INJUSTIFICADAS")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["BO"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["BP"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["BQ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["BR"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("TARDANZAS")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["BS"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["BT"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["BU"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["BV"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("Cumple puntualmente")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["BW"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["BX"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["BY"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["BZ"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("Envía correctamente")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["CA"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["CB"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["CC"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["CD"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("Justifica la inasistencia")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["CE"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["CF"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["CG"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["CH"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("Apoya en las tareas")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["CI"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["CJ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["CK"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["CL"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("Asiste a reuniones")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["CM"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["CN"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["CO"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["CP"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("Colabora con las")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["CQ"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["CR"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["CS"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["CT"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("APRECIACI")!= -1:
							if (n.PAcademico.Nombre.upper()=='1 BIMESTRE' or n.PAcademico.Nombre.upper()=='I BIMESTRE'):
								Hoja1["CU"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='2 BIMESTRE' or n.PAcademico.Nombre.upper()=='II BIMESTRE'):
								Hoja1["CV"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='3 BIMESTRE' or n.PAcademico.Nombre.upper()=='III BIMESTRE'):
								Hoja1["CW"+str(valor)] = n.Nota
							if (n.PAcademico.Nombre.upper()=='4 BIMESTRE' or n.PAcademico.Nombre.upper()=='IV BIMESTRE'):
								Hoja1["CX"+str(valor)] = n.Nota

						Hoja1["CY2"] = grado
		Libro.save(Ruta)
		#Aqui poner si es primaria o secundaria para que redirija a la descarga jeje
		if str(grado).find("PRIM")!= -1:
			return redirect("http://colcoopcv.com/static/files/RESUMEN_FINAL_PRIMARIA.xlsx")
		else:
			if str(grado).find("1SEC")!= -1 or str(grado).find("2SEC")!= -1 or str(grado).find("3SEC")!= -1:
				return redirect("http://colcoopcv.com/static/files/RESUMEN_FINAL_SECUNDARIA_1_2.xlsx")
			else:
				return redirect("http://colcoopcv.com/static/files/RESUMEN_FINAL_SECUNDARIA.xlsx")
	else:
		return render(request,'otras_opciones/imprimir_resumen_anual.html',contexto)

def ImprimirConsolidadoLibretas(request):
	aac = AnoAcademico.objects.all().order_by('-Ano')
	pac = PAcademico.objects.all().order_by('Nombre')
	contexto = {'aac':aac,'pac':pac}
	if request.method =='POST':
		grado=request.POST.get("Grado")
		seccion=request.POST.get("Seccion")

		ano=AnoAcademico()
		ano.id=request.POST.get("ano")

		paca=PAcademico()
		paca.id=request.POST.get("pacademico")

		#matri_notas = AvanceNotas.objects.filter(Matricula__Grado=grado,Matricula__Seccion=seccion,Matricula__AnoAcademico=ano,PAcademico=paca)
		matri = Matricula.objects.filter(Grado=grado,Seccion=seccion,AnoAcademico=ano).order_by('Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres')
		notas = Notas.objects.filter(Matricula__Grado=grado,Matricula__Seccion=seccion,Matricula__AnoAcademico=ano,PAcademico=paca)
		if str(grado).find("PRIM")!= -1:			
			Ruta = "static/files/PLANTILLA_LIBRETA_PRIMARIA.xlsx"
		else:
			Ruta = "static/files/PLANTILLA_LIBRETA_SECUNDARIA.xlsx"
		Libro = load_workbook(Ruta)
		Hoja1 = Libro.active
		#LIMPIANDO TODAS LAS CELDAS################
		Hoja1["B3"]=""#Limpiando PeriodoAcademico Grado y Seccion
		
		if str(grado).find("PRIM")!=-1:
			for filas in range(5,50):#LIMPIANDO_FILAS
				Hoja1["B"+str(filas)]=""
				for columnas in range(2,50):#LIMPIANDO_COLUMNAS
					Hoja1.cell(row=filas,column=columnas,value="")
		col=1
		if str(grado).find("SEC")!=-1:
			for filas in range(5,50):#LIMPIANDO_FILAS
				Hoja1["B"+str(filas)]=""
				#ESTA PENDIENTE ORDEN DE MERITO
				Hoja1["AB"+str(filas)]=""
				Hoja1["AC"+str(filas)]=""
				Hoja1["AD"+str(filas)]=""
				Hoja1["AE"+str(filas)]=""
				Hoja1["AF"+str(filas)]=""
				
				Hoja1["C"+str(filas)]=""
				Hoja1["E"+str(filas)]=""
				Hoja1["G"+str(filas)]=""
				Hoja1["I"+str(filas)]=""
				Hoja1["K"+str(filas)]=""
				Hoja1["M"+str(filas)]=""
				Hoja1["O"+str(filas)]=""
				Hoja1["Q"+str(filas)]=""
				Hoja1["S"+str(filas)]=""
				Hoja1["U"+str(filas)]=""
				Hoja1["W"+str(filas)]=""
				Hoja1["Y"+str(filas)]=""

				#for columnas in range(1,12):#LIMPIANDO_COLUMNAS
				#	col=col+2
				#	Hoja1.cell(row=filas,column=col,value="")
			##########################################################33
		#Si el ALumno tiene 2 notas el avance lo tomará la ultima nota
		valor=4
		for m in matri:
			valor = valor + 1
			Hoja1["B"+str(valor)] = m.Alumno.ApellidoPaterno + " " + m.Alumno.ApellidoMaterno + " " + m.Alumno.Nombres
			for n in notas:
				Hoja1["B3"]=n.PAcademico.Nombre.upper()+" - "+grado+" - "+seccion
				if (m.id==n.Matricula.id):
					if str(grado).find("SEC")!= -1:
						if str(n.Curso.Nombre).find("MATEMÁTICA")!= -1:
							Hoja1["C"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("COMUNICACIÓN")!= -1:
							Hoja1["E"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("INGLÉS")!= -1:
							Hoja1["G"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("CHINO MANDARÍN")!= -1:
							Hoja1["I"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("ARTE Y CULTURA")!= -1:
							Hoja1["K"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("CIENCIAS SOCIALES")!= -1:
							Hoja1["M"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("DESARROLLO PERSONAL")!= -1:
							Hoja1["O"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("EDUCACIÓN FÍSICA")!= -1:
							Hoja1["Q"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("EDUCACIÓN RELIGIOSA")!= -1:
							Hoja1["S"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("CIENCIA Y TECNOLOGÍA")!= -1:
							Hoja1["U"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("EDUCACIÓN PARA EL TRABAJO")!= -1:
							Hoja1["W"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("COMPORTAMIENTO")!= -1:
							Hoja1["Y"+str(valor)] = n.Nota
						#if str(n.Curso.Nombre).find("PUNTAJE")!= -1:
						#	Hoja1["Z"+str(valor)] = n.Nota
						#if str(n.Curso.Nombre).find("ORDEN DE MÉRITO")!= -1:
						#	Hoja1["AA"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("JUSTIFICADAS")!= -1:
							Hoja1["AB"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("INJUSTIFICADAS")!= -1:
							Hoja1["AC"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("TARDANZAS")!= -1:
							Hoja1["AD"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("APRECIACI")!= -1:
							Hoja1["AE"+str(valor)] = n.Nota
					else:#PRIMARIA
						if str(n.Curso.Nombre).find("MATEMÁTICA")!= -1:
							Hoja1["C"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("COMUNICACIÓN")!= -1:
							Hoja1["D"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("ARTE Y CULTURA")!= -1:
							Hoja1["E"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("PERSONAL SOCIAL")!= -1:
							Hoja1["F"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("EDUCACIÓN FÍSICA")!= -1:
							Hoja1["G"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("EDUCACIÓN RELIGIOSA")!= -1:
							Hoja1["H"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("CIENCIA Y TECNOLOGÍA")!= -1:
							Hoja1["I"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("INGLÉS")!= -1:
							Hoja1["J"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("CHINO MANDARÍN")!= -1:
							Hoja1["K"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("COMPUTACIÓN")!= -1:
							Hoja1["L"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("RESPETO")!= -1:
							Hoja1["M"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("PUNTUALIDAD")!= -1:
							Hoja1["N"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("ORDEN")!= -1:
							Hoja1["O"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("DISCIPLINA")!= -1:
							Hoja1["P"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("PARTICIPACIÓN")!= -1:
							Hoja1["Q"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("JUSTIFICADAS")!= -1:
							Hoja1["R"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("INJUSTIFICADAS")!= -1:
							Hoja1["S"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("TARDANZAS")!= -1:
							Hoja1["T"+str(valor)] = n.Nota
						#if str(n.Curso.Nombre).find("Cumple puntualmente")!= -1:
						#	Hoja1["U"+str(valor)] = n.Nota
						#if str(n.Curso.Nombre).find("Envía correctamente")!= -1:
						#	Hoja1["V"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("Justifica")!= -1:#cambiado3
							Hoja1["W"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("Apoya en las tareas")!= -1:##cambiado2
							Hoja1["X"+str(valor)] = n.Nota
						if str(n.Curso.Nombre).find("Asiste y participa")!= -1:##cambiado1
							Hoja1["Y"+str(valor)] = n.Nota
						#if str(n.Curso.Nombre).find("Colabora con las")!= -1:
						#	Hoja1["Z"+str(valor)] = n.Nota
						#if str(n.Curso.Nombre).find("APRECIACI")!= -1:
						#	Hoja1["AA"+str(valor)] = n.Nota

		Libro.save(Ruta)
		#Aqui poner si es primaria o secundaria para que redirija a la descarga jeje
		if str(grado).find("PRIM")!= -1:
			return redirect("http://colcoopcv.com/static/files/PLANTILLA_LIBRETA_PRIMARIA.xlsx")
		else:
			return redirect("http://colcoopcv.com/static/files/PLANTILLA_LIBRETA_SECUNDARIA.xlsx")
	else:
		return render(request,'otras_opciones/imprimir_consolidado_libretas.html',contexto)

def ImprimirConsolidadoAvances(request):
	aac = AnoAcademico.objects.all().order_by('-Ano')
	pac = PAcademico.objects.all().order_by('Nombre')
	contexto = {'aac':aac,'pac':pac}
	if request.method =='POST':
		grado=request.POST.get("Grado")
		seccion=request.POST.get("Seccion")

		ano=AnoAcademico()
		ano.id=request.POST.get("ano")

		paca=PAcademico()
		paca.id=request.POST.get("pacademico")

		#matri_notas = AvanceNotas.objects.filter(Matricula__Grado=grado,Matricula__Seccion=seccion,Matricula__AnoAcademico=ano,PAcademico=paca)
		matri = Matricula.objects.filter(Grado=grado,Seccion=seccion,AnoAcademico=ano).order_by('Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres')
		notas = AvanceNotasComp.objects.filter(Matricula__Grado=grado,Matricula__Seccion=seccion,Matricula__AnoAcademico=ano,PAcademico=paca)
		compecur=CompetenciaCurso.objects.all()
		if str(grado).find("PRIM")!= -1:
			Ruta = "static/files/PLANTILLA_AVANCE_PRIMARIA.xlsx"
		else:
			Ruta = "static/files/PLANTILLA_AVANCE_SECUNDARIA.xlsx"
		Libro = load_workbook(Ruta)
		Hoja1 = Libro.active
		##LIMPIANDO TODAS LAS CELDAS################3
		Hoja1["B3"]=""#Limpiando PeriodoAcademico Grado y Seccion
		for filas in range(6,50):#LIMPIANDO_FILAS
			Hoja1["B"+str(filas)]=""
			for columnas in range(2,50):#LIMPIANDO_COLUMNAS
				Hoja1.cell(row=filas,column=columnas,value="")
		##########################################################33
		#Si el ALumno tiene 2 notas el avance lo tomará la ultima nota
		valor=5
		col=2
		for m in matri:
			valor = valor + 1
			Hoja1["B"+str(valor)] = m.Alumno.ApellidoPaterno + " " + m.Alumno.ApellidoMaterno + " " + m.Alumno.Nombres
			for n in notas:
				Hoja1["B3"]=n.PAcademico.Nombre.upper()+" - "+grado+" - "+seccion
				if (m.id==n.Matricula.id):
					Hoja1["D"+str(valor)] = n.Nota
							# Hoja1["E"+str(valor)] = n.Nota
							# Hoja1["F"+str(valor)] = n.Nota
					# if str(grado).find("PRIM")!= -1: #PRIMARIA
					# 	if str(n.Curso.Nombre).find("MATEMÁTICA")!= -1:
					# 		Hoja1["C"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("COMUNICACIÓN")!= -1:
					# 		Hoja1["D"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("ARTE Y CULTURA")!= -1:
					# 		Hoja1["E"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("PERSONAL SOCIAL")!= -1:
					# 		Hoja1["F"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("EDUCACIÓN FÍSICA")!= -1:
					# 		Hoja1["G"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("EDUCACIÓN RELIGIOSA")!= -1:
					# 		Hoja1["H"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("CIENCIA Y TECNOLOGÍA")!= -1:
					# 		Hoja1["I"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("INGLÉS")!= -1:
					# 		Hoja1["J"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("CHINO MANDARÍN")!= -1:
					# 		Hoja1["K"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("COMPUTACIÓN")!= -1:
					# 		Hoja1["L"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("RESPETO")!= -1:
					# 		Hoja1["M"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("PUNTUALIDAD")!= -1:
					# 		Hoja1["N"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("ORDEN")!= -1:
					# 		Hoja1["O"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("DISCIPLINA")!= -1:
					# 		Hoja1["P"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("PARTICIPACIÓN")!= -1:
					# 		Hoja1["Q"+str(valor)] = n.Nota
					# else:#SECUNDARIA
					# 	if str(n.Curso.Nombre).find("MATEMÁTICA")!= -1:
					# 		Hoja1["C"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("RAZONAMIENTO MATEMÁTICO")!= -1:
					# 		Hoja1["D"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("COMUNICACIÓN")!= -1:
					# 		Hoja1["E"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("RAZONAMIENTO VERBAL")!= -1:
					# 		Hoja1["F"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("INGLÉS")!= -1:
					# 		Hoja1["G"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("CHINO MANDARÍN")!= -1:
					# 		Hoja1["H"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("ARTE Y CULTURA")!= -1:
					# 		Hoja1["I"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("CIENCIAS SOCIALES")!= -1:
					# 		Hoja1["J"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("DESARROLLO PERSONAL,")!= -1:
					# 		Hoja1["K"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("EDUCACIÓN FÍSICA")!= -1:
					# 		Hoja1["L"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("EDUCACIÓN RELIGIOSA")!= -1:
					# 		Hoja1["M"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("CIENCIA Y TECNOLOGÍA")!= -1:
					# 		Hoja1["N"+str(valor)] = n.Nota
					# 	if str(n.Curso.Nombre).find("EDUCACIÓN PARA EL TRABAJO")!= -1:
					# 		Hoja1["O"+str(valor)] = n.Nota
		Libro.save(Ruta)
		#Aqui poner si es primaria o secundaria para que redirija a la descarga jeje
		if str(grado).find("PRIM")!= -1:
			return redirect("/static/files/PLANTILLA_AVANCE_PRIMARIA.xlsx")
		else:
			return redirect("/static/files/PLANTILLA_AVANCE_SECUNDARIA.xlsx")
	else:
		return render(request,'otras_opciones/imprimir_consolidado_avances2.html',contexto)
###########################LIBRETAS##############################################

def ImprimirLibretas(request):
	#matri = Matricula() Comentado el  16-Jul-2019 9:44 p.m.
	#aac = AnoAcademico()
	aac = AnoAcademico.objects.all().order_by('-Ano')
	pac = PAcademico.objects.all().order_by('Nombre')
	contexto = {'aac':aac,'pac':pac}
	if request.method == 'POST':
		
		paca = PAcademico()
		paca.id = request.POST.get("Pac")
		
		ano = AnoAcademico() 
		ano.id = request.POST.get("Ano")

		gra = request.POST.get("Grado")
		sec = request.POST.get("Seccion")
		mat = Matricula.objects.filter(Grado=gra,Seccion=sec,AnoAcademico=ano).order_by('Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres')
		pdf_file_name=gra+'_'+sec+'.pdf'#nombre del archivo
		
		tutor = Docente.objects.get(TutorGrado=gra,TutorSeccion=sec)#Obtener Tutor	
		response = HttpResponse(content_type='application/pdf')
		response['Content-Disposition'] = 'inline; filename="'+pdf_file_name+'"'#del ejemplo

		buffer = BytesIO()#Del ejemplo
		c = canvas.Canvas(buffer, pagesize=portrait(A4))
		#paca = PAcademico.objects.get(FechaInicio__lt=datetime.now(),FechaFinal__gt=datetime.now())#Bimestre
		paca = PAcademico.objects.get(id=paca.id)#Bimestre
		
		for m in mat:#Para empezar a dibujar una libreta por cada matricula es una libreta
			
			c.drawImage("static/img/logo_cvallejo.png", 95, 715, width=55, height=70)
			c.drawImage("static/img/logo_goreloreto.png", 105, 790, width=80, height=35)
			c.drawImage("static/img/logo_minedu.png", 280, 780, width=75, height=45)
			c.drawImage("static/img/logo_dreloreto.png", 450, 780, width=50, height=50)
			c.drawImage("static/img/firma_segundo.png", 400, 30, width=130, height=55)#firma director encargado
			c.setFont('Helvetica-Bold', 12, leading=None)
			c.setFillColor(HexColor(0x4C9141))
			c.drawCentredString(300, 765, "COLEGIO COOPERATIVO CESAR VALLEJO")
			c.setFont('Helvetica-Bold', 11, leading=None)
			c.setFillColor(HexColor(0x0A0A0A))
			c.drawCentredString(300, 755, "NIVEL PRIMARIO-SECUNDARIO")
			c.setFont('Helvetica', 10, leading=None)	
			c.drawCentredString(300, 745, "Putumayo N° 966-Iquitos-Telef.973 891800")
			c.drawCentredString(300, 735, "Email: colegio_vallejo@yahoo.com")
			c.setFont('Helvetica-Bold', 12, leading=None)
			c.drawCentredString(300, 720, "LIBRETA DE NOTAS " + str(m.AnoAcademico))
			
			c.setFont('Helvetica', 9, leading=None)			
			
			bloq_cursos=60
			bloq_datos=35
			bloq_apre=-20
			if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC' or str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
				c.drawString(90, 705-bloq_datos, "ALUMNO (A):")
				c.drawString(90, 690-bloq_datos, "GRADO:")
				c.drawString(420, 705-bloq_datos,"NIVEL:")
				c.drawString(420, 690-bloq_datos,"BIMESTRE:")
				c.drawString(90, 675-bloq_datos, "TUTOR (A):")
				c.drawString(142, 675-bloq_datos, tutor.User.first_name + ' ' + tutor.User.last_name)
				c.roundRect(85, 667-bloq_datos,460,50,3)#Cuadro del Bloque Datos
			else:
				c.drawString(90, 705, "ALUMNO (A):")
				c.drawString(90, 690, "GRADO:")
				c.drawString(420, 705,"NIVEL:")
				c.drawString(420, 690,"BIMESTRE:")
				c.drawString(90, 675, "TUTOR (A):")
				c.drawString(142, 675, tutor.User.first_name + ' ' + tutor.User.last_name)
				c.roundRect(85, 667,460,50,3)#Cuadro del Bloque Datos

			w,h = A4#toma como referencia para hacer los rectangulos
			
			c.setFont('Helvetica',9, leading=None)
			grado=str(m.Grado)[0]#Extrae solo la primera letra
			
			
			
			nivel=str(m.Grado)[1:len(m.Grado)] #extrae solo PRIM
			if nivel=='PRIM':
				nivel='PRIMARIO'
			else:
				nivel='SECUNDARIO'
			

			if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC' or str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
				c.drawString(130, 690-bloq_datos, str(grado+'°'))
				c.drawString(145, 690-bloq_datos, str('"'+m.Seccion+'"'))
				c.drawString(150, 705-bloq_datos, str(m.Alumno.Nombres +' '+m.Alumno.ApellidoPaterno+' '+m.Alumno.ApellidoMaterno).upper())

				c.drawString(470, 690-bloq_datos, str(paca.Nombre).upper())#Periodo academico
				c.drawString(453, 705-bloq_datos,str(nivel).upper())
			else:
				c.drawString(130, 690, str(grado+'°'))
				c.drawString(145, 690, str('"'+m.Seccion+'"'))
				c.drawString(150, 705, str(m.Alumno.Nombres +' '+m.Alumno.ApellidoPaterno+' '+m.Alumno.ApellidoMaterno).upper())

				c.drawString(470, 690, str(paca.Nombre).upper())#Periodo academico
				c.drawString(453, 705,str(nivel).upper())


			noti = Notas.objects.filter(Matricula__id=m.id)
			#LIBRETA SECUNDARIA LIBRETA SECUNDARIA LIBRETA SECUNDARIA
			if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC' or str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':

				c.drawImage("static/img/firma_ruth.png", 90, 30, width=130, height=50)#firma ruth
				c.drawString(240, 630- bloq_cursos, "ÁREA")#Area
				c.drawString(436, 650- bloq_cursos, "BIMESTRES")#Bimestres
				c.drawString(424, 622- bloq_cursos, "I")#Bimestres
				c.drawString(446, 622- bloq_cursos, "II")#Bimestres
				c.drawString(468, 622- bloq_cursos, "III")#Bimestres
				c.drawString(487, 622- bloq_cursos, "IV")#Bimestres
				#se cambio el eje 
				c.roundRect(85, 612- bloq_cursos, 330, 50,0)#Rectangulo Izquierdo Area
				c.roundRect(415, 612- bloq_cursos, 22, 35,0)#Rectangulo I Bimestre
				c.roundRect(437, 612- bloq_cursos, 22, 35,0)#Rectangulo II Bimestre
				c.roundRect(459, 612- bloq_cursos, 22, 35,0)#Rectangulo III Bimestre
				c.roundRect(481, 612- bloq_cursos, 22, 35,0)#Rectangulo IV Bimestre
				c.roundRect(415, 647- bloq_cursos, 88, 15,0)#Rectangulo Bimestres
				c.drawString(505, 633- bloq_cursos, "Promedio")#Promedio
				c.drawString(513, 622- bloq_cursos, "Final")#Final
				c.roundRect(503, 612- bloq_cursos, 42, 50,0)#Rec PromedioFinal
				
				#Lineas de verticales de todos los cursos
				c.line(415,447- bloq_cursos,415,612- bloq_cursos)#x1 x2
				c.line(437,447- bloq_cursos,437,612- bloq_cursos)#x1 x2
				c.line(459,447- bloq_cursos,459,612- bloq_cursos)#x1 x2
				c.line(481,447- bloq_cursos,481,612- bloq_cursos)#x1 x2
				c.line(503,447- bloq_cursos,503,612- bloq_cursos)#x1 x2
				#Se comentaron para desapacer la parte de abajo de secundaria
				#Lineas de Comportamiento  orden de merito
				#c.line(415,417,415,387)#x1 x2
				#c.line(437,417,437,387)#x1 x2
				#c.line(459,417,459,387)#x1 x2
				#c.line(481,417,481,387)#x1 x2
				#c.line(503,417,503,387)#x1 x2
				#Lineas de Inasistencias
				#c.line(415,372,415,327)#x1 x2
				#c.line(437,372,437,327)#x1 x2
				#c.line(459,372,459,327)#x1 x2
				#c.line(481,372,481,327)#x1 x2
				#c.line(503,372,503,327)#x1 x2
				###########################################
				c.line(300,200,300,155)#x1 x2 LINEA PROMOVIDO

				c.roundRect(85,597- bloq_cursos, 460, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,582- bloq_cursos, 460, 15,0)#Rectangulo Cada Curso
				c.roundRect(85,567- bloq_cursos, 460, 15,0)#Rectangulo Cada Curso
				c.roundRect(85,552- bloq_cursos, 460, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,537- bloq_cursos, 460, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,522- bloq_cursos, 460, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,507- bloq_cursos, 460, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,492- bloq_cursos, 460, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,477- bloq_cursos, 460, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,462- bloq_cursos, 460, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,447- bloq_cursos, 460, 15,0)#Rectangulo Cada Curso						

				####SE COMENTO POR NO HABRA EVALUACION X LA CUARENTENA
				#c.roundRect(85,402, 460, 15,0)#Comportamiento
				#c.roundRect(85,387, 460, 15,0)#Orden de Mérito

				#c.roundRect(85,357, 460, 15,0)#Inasistencias Justificadas
				#c.roundRect(85,342, 460, 15,0)#Inasistencias Injustificadas
				#c.roundRect(85,327, 460, 15,0)#Tardanzas

				c.roundRect(85,297 - bloq_apre, 460, 15,0)#Apreciación del Tutor titulo
				c.roundRect(85,282 - bloq_apre, 460, 15,0)#I Bimestre
				c.roundRect(85,267 - bloq_apre, 460, 15,0)#II Bimestre 312,297,282,267
				c.roundRect(85,252 - bloq_apre, 460, 15,0)#III Bimestre
				c.roundRect(85,237 - bloq_apre, 460, 15,0)#IV Bimestre
				c.roundRect(150,200, 340, 15,0)#Resumen Final
				c.roundRect(150,185, 340, 15,0)#Promovido
				c.roundRect(150,170, 340, 15,0)#Pasa a Recuperación 
				c.roundRect(150,155, 340, 15,0)#Repite

				c.setFont('Helvetica-Bold', 9, leading=None)
				c.drawString(150, 301 - bloq_apre, "APRECIACIÓN DEL TUTOR(A) SOBRE LAS ACTITUDES DEL ALUMNO (A)")
				c.drawString(90, 286 - bloq_apre, "I BIMESTRE")
				c.drawString(90, 271 - bloq_apre, "II BIMESTRE")
				c.drawString(90, 256 - bloq_apre, "III BIMESTRE")
				c.drawString(90, 241 - bloq_apre, "IV BIMESTRE")
				c.drawString(285, 204, "RESUMEN FINAL")
				c.line(150,297- bloq_apre,150,237- bloq_apre)#

				c.setFont('Helvetica',9, leading=None)
				c.drawString(160, 189, "PROMOVIDO")
				c.drawString(160, 174, "PASA A RECUPERACIÓN")
				c.drawString(160, 159, "REPITE")
				
				
			#Desde aqui pasamos el for para generar cada rectangulo
				fila=h-230
				PRO_MAT=[]
				PRO_COMU=[]
				PRO_INGL=[]
				PRO_CHIN=[]
				PRO_ARTE=[]
				PRO_CIES=[]
				PRO_DPER=[]
				PRO_EFIS=[]
				PRO_RELI=[]
				PRO_CIEN=[]
				PRO_TRAB=[]
				PRO_GENERAL=[]
				
				for nn in noti:
					fila = fila - 15
					nombre_curso=str(nn.Curso.Nombre)[0:4]#extrae las 4 letras primera

					if nombre_curso=='MATE':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,600- bloq_cursos,nn.Curso.Nombre)#Escribiendo Curso Matematica
							MATE_IBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								MATE_IBIM=numeros_a_letras(MATE_IBIM)
								c.setFillColor(HexColor(letra_color(MATE_IBIM)))
								c.drawString(419,600- bloq_cursos, MATE_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_MAT.append(MATE_IBIM)
								c.setFillColor(HexColor(pone_color(MATE_IBIM)))
								c.drawString(419,600- bloq_cursos, MATE_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,600- bloq_cursos,nn.Curso.Nombre)
							MATE_IIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								MATE_IIBIM=numeros_a_letras(MATE_IIBIM)
								c.setFillColor(HexColor(letra_color(MATE_IIBIM)))
								c.drawString(441,600- bloq_cursos,MATE_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_MAT.append(MATE_IIBIM)
								c.setFillColor(HexColor(pone_color(MATE_IIBIM)))
								c.drawString(441,600- bloq_cursos,MATE_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,600- bloq_cursos,nn.Curso.Nombre)
							MATE_IIIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								MATE_IIIBIM=numeros_a_letras(MATE_IIIBIM)
								c.setFillColor(HexColor(letra_color(MATE_IIIBIM)))
								c.drawString(463,600- bloq_cursos,MATE_IIIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_MAT.append(MATE_IIIBIM)
								c.setFillColor(HexColor(pone_color(MATE_IIIBIM)))
								c.drawString(463,600- bloq_cursos,MATE_IIIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,600- bloq_cursos,nn.Curso.Nombre)
							MATE_IVBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								PRO_GENERAL.append(MATE_IVBIM)
								MATE_IVBIM=numeros_a_letras(MATE_IVBIM)
								c.setFillColor(HexColor(letra_color(MATE_IVBIM)))
								c.drawString(485,600- bloq_cursos,MATE_IVBIM)#
								#Imprimiendo Promedio
								c.drawString(517,600- bloq_cursos,MATE_IVBIM)#se agregó_15Dic2020, esto imprime en promedio final
								c.setFillColor(HexColor(0x0A0A0A))#bicado en la columan II Bimestre
							else:
								PRO_MAT.append(MATE_IVBIM)
								
								c.setFillColor(HexColor(pone_color(MATE_IVBIM)))
								c.drawString(485,600- bloq_cursos,MATE_IVBIM)
								c.setFillColor(HexColor(0x0A0A0A))
						
						if len(PRO_MAT)==4:
							promedio_mat=calc_promedio(PRO_MAT)
							c.setFillColor(HexColor(pone_color(promedio_mat)))
							c.drawString(517, 600- bloq_cursos,cero_izq(promedio_mat))
							c.setFillColor(HexColor(0x0A0A0A))
							PRO_GENERAL.append(promedio_mat)

					if nombre_curso=='COMU':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':
							c.drawString(90,585- bloq_cursos,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							COMU_IBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								COMU_IBIM=numeros_a_letras(COMU_IBIM)
								c.setFillColor(HexColor(letra_color(COMU_IBIM)))
								c.drawString(419, 585- bloq_cursos,COMU_IBIM)
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_COMU.append(COMU_IBIM)
								c.setFillColor(HexColor(pone_color(COMU_IBIM)))
								c.drawString(419, 585- bloq_cursos,COMU_IBIM)
								c.setFillColor(HexColor(0x0A0A0A))

						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,585- bloq_cursos,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							COMU_IIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								COMU_IIBIM=numeros_a_letras(COMU_IIBIM)
								c.setFillColor(HexColor(letra_color(COMU_IIBIM)))
								c.drawString(441, 585- bloq_cursos, COMU_IIBIM)#Ubicado en la columan II Bimestre		
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_COMU.append(COMU_IIBIM)
								c.setFillColor(HexColor(pone_color(COMU_IIBIM)))
								c.drawString(441, 585- bloq_cursos, COMU_IIBIM)#Ubicado en la columan II Bimestre		
								c.setFillColor(HexColor(0x0A0A0A))

						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,585- bloq_cursos,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							COMU_IIIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								COMU_IIIBIM=numeros_a_letras(COMU_IIIBIM)
								c.setFillColor(HexColor(letra_color(COMU_IIIBIM)))
								c.drawString(463, 585- bloq_cursos, COMU_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_COMU.append(COMU_IIIBIM)
								c.setFillColor(HexColor(pone_color(COMU_IIIBIM)))
								c.drawString(463, 585- bloq_cursos, COMU_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,585- bloq_cursos,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							COMU_IVBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								PRO_GENERAL.append(COMU_IVBIM)
								COMU_IVBIM=numeros_a_letras(COMU_IVBIM)
								c.setFillColor(HexColor(letra_color(COMU_IVBIM)))
								c.drawString(485, 585- bloq_cursos, COMU_IVBIM)#Ubicado en la columan IV Bimestre
								#imprimiendo promedio
								c.drawString(517, 585- bloq_cursos,COMU_IVBIM)
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_COMU.append(COMU_IVBIM)
								c.setFillColor(HexColor(pone_color(COMU_IVBIM)))
								c.drawString(485, 585- bloq_cursos, COMU_IVBIM)#Ubicado en la columan IV Bimestre
								c.setFillColor(HexColor(0x0A0A0A))

						if len(PRO_COMU)==4:
							promedio_com=calc_promedio(PRO_COMU)
							c.setFillColor(HexColor(pone_color(promedio_com)))
							c.drawString(517, 585- bloq_cursos,cero_izq(promedio_com))
							c.setFillColor(HexColor(0x0A0A0A))
							PRO_GENERAL.append(promedio_com)

					if nombre_curso=='INGL':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,570- bloq_cursos,nn.Curso.Nombre)
							INGL_IBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								INGL_IBIM=numeros_a_letras(INGL_IBIM)
								c.setFillColor(HexColor(letra_color(INGL_IBIM)))
								c.drawString(419, 570- bloq_cursos,INGL_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_INGL.append(INGL_IBIM)
								c.setFillColor(HexColor(pone_color(INGL_IBIM)))
								c.drawString(419, 570- bloq_cursos,INGL_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,570- bloq_cursos,nn.Curso.Nombre)
							INGL_IIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								INGL_IIBIM=numeros_a_letras(INGL_IIBIM)
								c.setFillColor(HexColor(letra_color(INGL_IIBIM)))
								c.drawString(441, 570- bloq_cursos, INGL_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_INGL.append(INGL_IIBIM)
								c.setFillColor(HexColor(pone_color(INGL_IIBIM)))
								c.drawString(441, 570- bloq_cursos, INGL_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,570- bloq_cursos,nn.Curso.Nombre)
							INGL_IIIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								INGL_IIIBIM=numeros_a_letras(INGL_IIIBIM)
								c.setFillColor(HexColor(letra_color(INGL_IIIBIM)))
								c.drawString(463, 570- bloq_cursos, INGL_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_INGL.append(INGL_IIIBIM)
								c.setFillColor(HexColor(pone_color(INGL_IIIBIM)))
								c.drawString(463, 570- bloq_cursos, INGL_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))

						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,570- bloq_cursos,nn.Curso.Nombre)
							INGL_IVBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								PRO_GENERAL.append(INGL_IVBIM)
								INGL_IVBIM=numeros_a_letras(INGL_IVBIM)
								c.setFillColor(HexColor(letra_color(INGL_IVBIM)))
								c.drawString(485, 570- bloq_cursos, INGL_IVBIM)#Ubicado en la columan IV Bimestre
								#imprimiendo promedio final
								c.drawString(517, 570- bloq_cursos,INGL_IVBIM)
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_INGL.append(INGL_IVBIM)
								c.setFillColor(HexColor(pone_color(INGL_IVBIM)))
								c.drawString(485, 570- bloq_cursos, INGL_IVBIM)#Ubicado en la columan IV Bimestre
								c.setFillColor(HexColor(0x0A0A0A))

						if len(PRO_INGL)==4:
							promedio_ingl=calc_promedio(PRO_INGL)
							c.setFillColor(HexColor(pone_color(promedio_ingl)))
							c.drawString(517, 570- bloq_cursos,cero_izq(promedio_ingl))
							c.setFillColor(HexColor(0x0A0A0A))
							PRO_GENERAL.append(promedio_ingl)

					if nombre_curso=='CHIN':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,555- bloq_cursos,nn.Curso.Nombre)
							CHIN_IBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								CHIN_IBIM=numeros_a_letras(CHIN_IBIM)
								c.setFillColor(HexColor(letra_color(CHIN_IBIM)))
								c.drawString(419, 555- bloq_cursos,CHIN_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_CHIN.append(CHIN_IBIM)
								c.setFillColor(HexColor(pone_color(CHIN_IBIM)))
								c.drawString(419, 555- bloq_cursos,CHIN_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,555- bloq_cursos,nn.Curso.Nombre)
							CHIN_IIBIM=nn.Nota

							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':	
								CHIN_IIBIM=numeros_a_letras(CHIN_IIBIM)
								c.setFillColor(HexColor(letra_color(CHIN_IIBIM)))
								c.drawString(441, 555- bloq_cursos, CHIN_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_CHIN.append(CHIN_IIBIM)
								c.setFillColor(HexColor(pone_color(CHIN_IIBIM)))
								c.drawString(441, 555- bloq_cursos, CHIN_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,555- bloq_cursos,nn.Curso.Nombre)
							CHIN_IIIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':		
								CHIN_IIIBIM=numeros_a_letras(CHIN_IIIBIM)
								c.setFillColor(HexColor(letra_color(CHIN_IIIBIM)))
								c.drawString(463, 555- bloq_cursos, CHIN_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_CHIN.append(CHIN_IIIBIM)
								c.setFillColor(HexColor(pone_color(CHIN_IIIBIM)))
								c.drawString(463, 555- bloq_cursos, CHIN_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,555- bloq_cursos,nn.Curso.Nombre)
							CHIN_IVBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								PRO_GENERAL.append(CHIN_IVBIM)
								CHIN_IVBIM=numeros_a_letras(CHIN_IVBIM)
								c.setFillColor(HexColor(letra_color(CHIN_IVBIM)))
								c.drawString(485, 555- bloq_cursos, CHIN_IVBIM)#Ubicado en la columan IV Bimestre
								#imprimiendo promedio final
								c.drawString(517, 555- bloq_cursos,CHIN_IVBIM)
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_CHIN.append(CHIN_IVBIM)
								c.setFillColor(HexColor(pone_color(CHIN_IVBIM)))
								c.drawString(485, 555- bloq_cursos, CHIN_IVBIM)#Ubicado en la columan IV Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						
						if len(PRO_CHIN)==4:
							promedio_chin=calc_promedio(PRO_CHIN)
							c.setFillColor(HexColor(pone_color(promedio_chin)))
							c.drawString(517, 555- bloq_cursos,cero_izq(promedio_chin))
							c.setFillColor(HexColor(0x0A0A0A))
							PRO_GENERAL.append(promedio_chin)

					if nombre_curso=='ARTE':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,540- bloq_cursos,nn.Curso.Nombre)
							ARTE_IBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								ARTE_IBIM=numeros_a_letras(ARTE_IBIM)
								c.setFillColor(HexColor(letra_color(ARTE_IBIM)))
								c.drawString(419, 540- bloq_cursos,ARTE_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_ARTE.append(ARTE_IBIM)
								c.setFillColor(HexColor(pone_color(ARTE_IBIM)))
								c.drawString(419, 540- bloq_cursos,ARTE_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,540- bloq_cursos,nn.Curso.Nombre)
							ARTE_IIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								ARTE_IIBIM=numeros_a_letras(ARTE_IIBIM)
								c.setFillColor(HexColor(letra_color(ARTE_IIBIM)))
								c.drawString(441, 540- bloq_cursos, ARTE_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_ARTE.append(ARTE_IIBIM)
								c.setFillColor(HexColor(pone_color(ARTE_IIBIM)))
								c.drawString(441, 540- bloq_cursos, ARTE_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,540- bloq_cursos,nn.Curso.Nombre)
							ARTE_IIIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								ARTE_IIIBIM=numeros_a_letras(ARTE_IIIBIM)
								c.setFillColor(HexColor(letra_color(ARTE_IIIBIM)))
								c.drawString(463, 540- bloq_cursos, ARTE_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_ARTE.append(ARTE_IIIBIM)
								c.setFillColor(HexColor(pone_color(ARTE_IIIBIM)))
								c.drawString(463, 540- bloq_cursos, ARTE_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,540- bloq_cursos,nn.Curso.Nombre)
							ARTE_IVBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								PRO_GENERAL.append(ARTE_IVBIM)
								ARTE_IVBIM=numeros_a_letras(ARTE_IVBIM)
								c.setFillColor(HexColor(letra_color(ARTE_IVBIM)))
								c.drawString(485, 540- bloq_cursos, ARTE_IVBIM)#Ubicado en la columan IV Bimestre
								#imprimiendo promedio final
								c.drawString(517, 540- bloq_cursos,ARTE_IVBIM)
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_ARTE.append(ARTE_IVBIM)
								c.setFillColor(HexColor(pone_color(ARTE_IVBIM)))
								c.drawString(485, 540- bloq_cursos, ARTE_IVBIM)#Ubicado en la columan IV Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						
						if len(PRO_ARTE)==4:
							promedio_art=calc_promedio(PRO_ARTE)
							c.setFillColor(HexColor(pone_color(promedio_art)))
							c.drawString(517, 540- bloq_cursos,cero_izq(promedio_art))
							c.setFillColor(HexColor(0x0A0A0A))
							PRO_GENERAL.append(promedio_art)

					if str(nn.Curso.Nombre).find("CIENCIAS SOCIALES")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,525- bloq_cursos,nn.Curso.Nombre)
							CIES_IBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								CIES_IBIM=numeros_a_letras(CIES_IBIM)
								c.setFillColor(HexColor(letra_color(CIES_IBIM)))
								c.drawString(419, 525- bloq_cursos,CIES_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_CIES.append(CIES_IBIM)
								c.setFillColor(HexColor(pone_color(CIES_IBIM)))
								c.drawString(419, 525- bloq_cursos,CIES_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,525- bloq_cursos,nn.Curso.Nombre)
							CIES_IIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								CIES_IIBIM=numeros_a_letras(CIES_IIBIM)
								c.setFillColor(HexColor(letra_color(CIES_IIBIM)))
								c.drawString(441, 525- bloq_cursos, CIES_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_CIES.append(CIES_IIBIM)
								c.setFillColor(HexColor(pone_color(CIES_IIBIM)))
								c.drawString(441, 525- bloq_cursos, CIES_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,525- bloq_cursos,nn.Curso.Nombre)
							CIES_IIIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								CIES_IIIBIM=numeros_a_letras(CIES_IIIBIM)
								c.setFillColor(HexColor(letra_color(CIES_IIIBIM)))
								c.drawString(463, 525- bloq_cursos, CIES_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_CIES.append(CIES_IIIBIM)
								c.setFillColor(HexColor(pone_color(CIES_IIIBIM)))
								c.drawString(463, 525- bloq_cursos, CIES_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,525- bloq_cursos,nn.Curso.Nombre)
							CIES_IVBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								PRO_GENERAL.append(CIES_IVBIM)
								CIES_IVBIM=numeros_a_letras(CIES_IVBIM)
								c.setFillColor(HexColor(letra_color(CIES_IVBIM)))
								c.drawString(485, 525- bloq_cursos, CIES_IVBIM)#Ubicado en la columan IV Bimestre
								#imprimiendo promedio final
								c.drawString(517,525- bloq_cursos,CIES_IVBIM)
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_CIES.append(CIES_IVBIM)
								c.setFillColor(HexColor(pone_color(CIES_IVBIM)))
								c.drawString(485, 525- bloq_cursos, CIES_IVBIM)#Ubicado en la columan IV Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						
						if len(PRO_CIES)==4:
							promedio_cies=calc_promedio(PRO_CIES)
							c.setFillColor(HexColor(pone_color(promedio_cies)))
							c.drawString(517, 525- bloq_cursos,cero_izq(promedio_cies))
							c.setFillColor(HexColor(0x0A0A0A))
							PRO_GENERAL.append(promedio_cies)

					if str(nn.Curso.Nombre).find("DESARROLLO PERSONAL")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,510- bloq_cursos,nn.Curso.Nombre)
							DPER_IBIM=nn.Nota
						
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':		
								DPER_IBIM=numeros_a_letras(DPER_IBIM)
								c.setFillColor(HexColor(letra_color(DPER_IBIM)))	
								c.drawString(419, 510- bloq_cursos,DPER_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_DPER.append(DPER_IBIM)
								c.setFillColor(HexColor(pone_color(DPER_IBIM)))	
								c.drawString(419, 510- bloq_cursos,DPER_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,510- bloq_cursos,nn.Curso.Nombre)
							DPER_IIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								DPER_IIBIM=numeros_a_letras(DPER_IIBIM)
								c.setFillColor(HexColor(letra_color(DPER_IIBIM)))	
								c.drawString(441, 510- bloq_cursos, DPER_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_DPER.append(DPER_IIBIM)
								c.setFillColor(HexColor(pone_color(DPER_IIBIM)))	
								c.drawString(441, 510- bloq_cursos, DPER_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))

						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,510- bloq_cursos,nn.Curso.Nombre)
							DPER_IIIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								DPER_IIIBIM=numeros_a_letras(DPER_IIIBIM)
								c.setFillColor(HexColor(letra_color(DPER_IIIBIM)))	
								c.drawString(463, 510- bloq_cursos, DPER_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_DPER.append(DPER_IIIBIM)
								c.setFillColor(HexColor(pone_color(DPER_IIIBIM)))	
								c.drawString(463, 510- bloq_cursos, DPER_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,510- bloq_cursos,nn.Curso.Nombre)
							DPER_IVBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								PRO_GENERAL.append(DPER_IVBIM)
								DPER_IVBIM=numeros_a_letras(DPER_IVBIM)
								c.setFillColor(HexColor(letra_color(DPER_IVBIM)))	
								c.drawString(485, 510- bloq_cursos, DPER_IVBIM)#Ubicado en la columan IV Bimestre
								#imprimiendo promedio final
								c.drawString(517, 510- bloq_cursos,DPER_IVBIM)
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_DPER.append(DPER_IVBIM)
								c.setFillColor(HexColor(pone_color(DPER_IVBIM)))	
								c.drawString(485, 510- bloq_cursos, DPER_IVBIM)#Ubicado en la columan IV Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if len(PRO_DPER)==4:
							promedio_dper=calc_promedio(PRO_DPER)
							c.setFillColor(HexColor(pone_color(promedio_dper)))
							c.drawString(517, 510- bloq_cursos,cero_izq(promedio_dper))
							c.setFillColor(HexColor(0x0A0A0A))
							PRO_GENERAL.append(promedio_dper)

					if str(nn.Curso.Nombre).find("FÍSICA")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,495- bloq_cursos,nn.Curso.Nombre)
							EFIS_IBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								EFIS_IBIM=numeros_a_letras(EFIS_IBIM)
								c.setFillColor(HexColor(letra_color(EFIS_IBIM)))	
								c.drawString(419, 495- bloq_cursos,EFIS_IBIM)#ubicado en la columna IBimestre
							else:
								PRO_EFIS.append(EFIS_IBIM)
								c.setFillColor(HexColor(pone_color(EFIS_IBIM)))	
								c.drawString(419, 495- bloq_cursos,EFIS_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,495- bloq_cursos,nn.Curso.Nombre)
							EFIS_IIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								EFIS_IIBIM=numeros_a_letras(EFIS_IIBIM)
								c.setFillColor(HexColor(letra_color(EFIS_IIBIM)))	
								c.drawString(441, 495- bloq_cursos, EFIS_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_EFIS.append(EFIS_IIBIM)
								c.setFillColor(HexColor(pone_color(EFIS_IIBIM)))	
								c.drawString(441, 495- bloq_cursos, EFIS_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,495- bloq_cursos,nn.Curso.Nombre)
							EFIS_IIIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								EFIS_IIIBIM=numeros_a_letras(EFIS_IIIBIM)
								c.setFillColor(HexColor(letra_color(EFIS_IIIBIM)))	
								c.drawString(463, 495- bloq_cursos, EFIS_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_EFIS.append(EFIS_IIIBIM)
								c.setFillColor(HexColor(pone_color(EFIS_IIIBIM)))	
								c.drawString(463, 495- bloq_cursos, EFIS_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,495- bloq_cursos,nn.Curso.Nombre)
							EFIS_IVBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								PRO_GENERAL.append(EFIS_IVBIM)
								EFIS_IVBIM=numeros_a_letras(EFIS_IVBIM)
								c.setFillColor(HexColor(letra_color(EFIS_IVBIM)))	
								c.drawString(485, 495- bloq_cursos, EFIS_IVBIM)#Ubicado en la columan IV Bimestre
								#imprimiendo promedio final
								c.drawString(517, 495- bloq_cursos,EFIS_IVBIM)
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_EFIS.append(EFIS_IVBIM)
								c.setFillColor(HexColor(pone_color(EFIS_IVBIM)))	
								c.drawString(485, 495- bloq_cursos, EFIS_IVBIM)#Ubicado en la columan IV Bimestre
								c.setFillColor(HexColor(0x0A0A0A))

						if len(PRO_EFIS)==4:
							promedio_efis=calc_promedio(PRO_EFIS)
							c.setFillColor(HexColor(pone_color(promedio_efis)))
							c.drawString(517, 495- bloq_cursos,cero_izq(promedio_efis))
							c.setFillColor(HexColor(0x0A0A0A))
							PRO_GENERAL.append(promedio_efis)

					if str(nn.Curso.Nombre).find("RELIGI")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,480- bloq_cursos,nn.Curso.Nombre)
							RELI_IBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								RELI_IBIM=numeros_a_letras(RELI_IBIM)
								c.setFillColor(HexColor(letra_color(RELI_IBIM)))	
								c.drawString(419, 480- bloq_cursos,RELI_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_RELI.append(RELI_IBIM)
								c.setFillColor(HexColor(pone_color(RELI_IBIM)))	
								c.drawString(419, 480- bloq_cursos,RELI_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,480- bloq_cursos,nn.Curso.Nombre)
							RELI_IIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								RELI_IIBIM=numeros_a_letras(RELI_IIBIM)
								c.setFillColor(HexColor(letra_color(RELI_IIBIM)))
								c.drawString(441, 480- bloq_cursos, RELI_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_RELI.append(RELI_IIBIM)
								c.setFillColor(HexColor(pone_color(RELI_IIBIM)))
								c.drawString(441, 480- bloq_cursos, RELI_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,480- bloq_cursos,nn.Curso.Nombre)
							RELI_IIIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								RELI_IIIBIM=numeros_a_letras(RELI_IIIBIM)
								c.setFillColor(HexColor(letra_color(RELI_IIIBIM)))
								c.drawString(463, 480- bloq_cursos, RELI_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_RELI.append(RELI_IIIBIM)
								c.setFillColor(HexColor(pone_color(RELI_IIIBIM)))
								c.drawString(463, 480- bloq_cursos, RELI_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,480- bloq_cursos,nn.Curso.Nombre)
							RELI_IVBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								PRO_GENERAL.append(RELI_IVBIM)
								RELI_IVBIM=numeros_a_letras(RELI_IVBIM)
								c.setFillColor(HexColor(letra_color(RELI_IVBIM)))
								c.drawString(485, 480- bloq_cursos, RELI_IVBIM)#Ubicado en la columan IV Bimestre
								#imprimiendo promediofinal
								c.drawString(517, 480- bloq_cursos,RELI_IVBIM)
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_RELI.append(RELI_IVBIM)
								c.setFillColor(HexColor(pone_color(RELI_IVBIM)))
								c.drawString(485, 480- bloq_cursos, RELI_IVBIM)#Ubicado en la columan IV Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						
						if len(PRO_RELI)==4:
							promedio_reli=calc_promedio(PRO_RELI)
							c.setFillColor(HexColor(pone_color(promedio_reli)))
							c.drawString(517, 480- bloq_cursos,cero_izq(promedio_reli))
							c.setFillColor(HexColor(0x0A0A0A))
							PRO_GENERAL.append(promedio_reli)

					if str(nn.Curso.Nombre).find("CIENCIA Y TECNOLOG")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,465- bloq_cursos,nn.Curso.Nombre)
							CIEN_IBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								CIEN_IBIM=numeros_a_letras(CIEN_IBIM)
								c.setFillColor(HexColor(letra_color(CIEN_IBIM)))
								c.drawString(419, 465- bloq_cursos,CIEN_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_CIEN.append(CIEN_IBIM)
								c.setFillColor(HexColor(pone_color(CIEN_IBIM)))
								c.drawString(419, 465- bloq_cursos,CIEN_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,465- bloq_cursos,nn.Curso.Nombre)
							CIEN_IIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								CIEN_IIBIM=numeros_a_letras(CIEN_IIBIM)
								c.setFillColor(HexColor(letra_color(CIEN_IIBIM)))
								c.drawString(441, 465- bloq_cursos, CIEN_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_CIEN.append(CIEN_IIBIM)
								c.setFillColor(HexColor(pone_color(CIEN_IIBIM)))
								c.drawString(441, 465- bloq_cursos, CIEN_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,465- bloq_cursos,nn.Curso.Nombre)
							CIEN_IIIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								CIEN_IIIBIM=numeros_a_letras(CIEN_IIIBIM)
								c.setFillColor(HexColor(letra_color(CIEN_IIIBIM)))
								c.drawString(463, 465- bloq_cursos, CIEN_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_CIEN.append(CIEN_IIIBIM)
								c.setFillColor(HexColor(pone_color(CIEN_IIIBIM)))
								c.drawString(463, 465- bloq_cursos, CIEN_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,465- bloq_cursos,nn.Curso.Nombre)
							CIEN_IVBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								PRO_GENERAL.append(CIEN_IVBIM)
								CIEN_IVBIM=numeros_a_letras(CIEN_IVBIM)
								c.setFillColor(HexColor(letra_color(CIEN_IVBIM)))
								c.drawString(485, 465- bloq_cursos, CIEN_IVBIM)#Ubicado en la columan IV Bimestre
								#imprimiendo promedio final
								c.drawString(517, 465- bloq_cursos,CIEN_IVBIM)
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_CIEN.append(CIEN_IVBIM)
								c.setFillColor(HexColor(pone_color(CIEN_IVBIM)))
								c.drawString(485, 465- bloq_cursos, CIEN_IVBIM)#Ubicado en la columan IV Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						
						if len(PRO_CIEN)==4:
							promedio_cien=calc_promedio(PRO_CIEN)
							c.setFillColor(HexColor(pone_color(promedio_cien)))
							c.drawString(517, 465- bloq_cursos,cero_izq(promedio_cien))
							c.setFillColor(HexColor(0x0A0A0A))
							PRO_GENERAL.append(promedio_cien)

					if str(nn.Curso.Nombre).find("TRABAJO")!= -1:
						
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,450- bloq_cursos,nn.Curso.Nombre)
							TRAB_IBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								TRAB_IBIM=numeros_a_letras(TRAB_IBIM)
								c.setFillColor(HexColor(letra_color(TRAB_IBIM)))
								c.drawString(419, 450- bloq_cursos,TRAB_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_TRAB.append(TRAB_IBIM)
								c.setFillColor(HexColor(pone_color(TRAB_IBIM)))
								c.drawString(419, 450- bloq_cursos,TRAB_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,450- bloq_cursos,nn.Curso.Nombre)
							TRAB_IIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								TRAB_IIBIM=numeros_a_letras(TRAB_IIBIM)
								c.setFillColor(HexColor(letra_color(TRAB_IIBIM)))
								c.drawString(441, 450- bloq_cursos, TRAB_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_TRAB.append(TRAB_IIBIM)
								c.setFillColor(HexColor(pone_color(TRAB_IIBIM)))
								c.drawString(441, 450- bloq_cursos, TRAB_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,450- bloq_cursos,nn.Curso.Nombre)
							TRAB_IIIBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								TRAB_IIIBIM=numeros_a_letras(TRAB_IIIBIM)
								c.setFillColor(HexColor(letra_color(TRAB_IIIBIM)))
								c.drawString(463, 450- bloq_cursos, TRAB_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								PRO_TRAB.append(TRAB_IIIBIM)
								c.setFillColor(HexColor(pone_color(TRAB_IIIBIM)))
								c.drawString(463, 450- bloq_cursos, TRAB_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,450- bloq_cursos,nn.Curso.Nombre)
							TRAB_IVBIM=nn.Nota
							
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								PRO_GENERAL.append(TRAB_IVBIM)
								TRAB_IVBIM=numeros_a_letras(TRAB_IVBIM)
								c.setFillColor(HexColor(letra_color(TRAB_IVBIM)))
								c.drawString(485, 450- bloq_cursos, TRAB_IVBIM)#Ubicado en la columan IV Bimestre
								#imprimiendo promedio final
								c.drawString(517, 450- bloq_cursos,TRAB_IVBIM)
								c.setFillColor(HexColor(0x0A0A0A))
							else:####todo el ese imprimi el 4 bimestre de 4 y 5 sec
								PRO_TRAB.append(TRAB_IVBIM)
								c.setFillColor(HexColor(pone_color(TRAB_IVBIM)))
								c.drawString(485, 450- bloq_cursos, TRAB_IVBIM)
								c.setFillColor(HexColor(0x0A0A0A))

						
						#PRO_TRAB=calc_promedio(PRO_TRAB)
						PRO_GENERAL.append(PRO_TRAB)#para saber su estado situacional
						#c.setFillColor(HexColor(pone_color(PRO_TRAB)))
						#ff=0
						#for x in PRO_TRAB:
						#	ff+=1						
						
						if len(PRO_TRAB)==4:
							promedio_trab=calc_promedio(PRO_TRAB)
							c.setFillColor(HexColor(pone_color(promedio_trab)))
							c.drawString(517, 450- bloq_cursos,cero_izq(promedio_trab))
							c.setFillColor(HexColor(0x0A0A0A))
							PRO_GENERAL.append(promedio_trab)
						#c.drawString(517, 450- bloq_cursos,cero_izq(TRAB_IBIM))
						c.setFillColor(HexColor(0x0A0A0A))
						#PRO_TRAB=[]

					if str(nn.Curso.Nombre).find("COMPORTA")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,405- bloq_cursos,nn.Curso.Nombre)
							COMP_IBIM=nn.Nota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								COMP_IBIM_numeros_a_letras(COMP_IBIM)
								c.setFillColor(HexColor(letra_color(COMP_IBIM)))
								c.drawString(419, 405- bloq_cursos,COMP_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(letra_color(COMP_IBIM)))
								c.drawString(419, 405- bloq_cursos,COMP_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))

						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,405- bloq_cursos,nn.Curso.Nombre)
							COMP_IIBIM=nn.Nota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								
								COMP_IIBIM=numeros_a_letras(COMP_IIBIM)
								c.setFillColor(HexColor(letra_color(COMP_IIBIM)))
								c.drawString(441, 405- bloq_cursos, COMP_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(letra_color(COMP_IIBIM)))
								c.drawString(441, 405- bloq_cursos, COMP_IIBIM)#Ubicado en la columan II Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,405- bloq_cursos,nn.Curso.Nombre)
							COMP_IIIBIM=nn.Nota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								COMP_IIIBIM=numeros_a_letras(COMP_IIIBIM)
								c.setFillColor(HexColor(letra_color(COMP_IIIBIM)))
								c.drawString(463, 405- bloq_cursos, COMP_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(letra_color(COMP_IIIBIM)))
								c.drawString(463, 405- bloq_cursos, COMP_IIIBIM)#Ubicado en la columan III Bimestre
								c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,405- bloq_cursos,nn.Curso.Nombre)
							COMP_IVBIM=nn.Nota#COMPORTAMIENTO ES PURA LETRA
							c.setFillColor(HexColor(letra_color(COMP_IVBIM)))
							c.drawString(485, 405- bloq_cursos, COMP_IVBIM)
							c.setFillColor(HexColor(0x0A0A0A))
							
							if str(COMP_IBIM)!='' and str(COMP_IIBIM)!='' and str(COMP_IIIBIM)!='' and str(COMP_IVBIM)!='':
								c.drawString(520, 405- bloq_cursos,COMP_IVBIM)

					if str(nn.Curso.Nombre).find("ORDEN DE MÉRITO")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,390,nn.Curso.Nombre)
							ORDE_IBIM=nn.Nota
							c.drawString(419, 390,ORDE_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,390,nn.Curso.Nombre)
							ORDE_IIBIM=nn.Nota
							c.drawString(441, 390, ORDE_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,390,nn.Curso.Nombre)
							ORDE_IIIBIM=nn.Nota
							c.drawString(463, 390, ORDE_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,390,nn.Curso.Nombre)
							ORDE_IVBIM=nn.Nota
							c.drawString(485, 390, ORDE_IVBIM)#Ubicado en la columan IV Bimestre

					if str(nn.Curso.Nombre).find("PUESTO FINAL")!= -1:
						c.drawString(517, 390,nn.Nota)

					if nombre_curso=='JUST':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,360,'INASISTENCIAS '+nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							JUST_IBIM=nn.Nota
							c.drawString(419, 360,JUST_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,360,'INASISTENCIAS '+nn.Curso.Nombre)
							JUST_IIBIM=nn.Nota
							c.drawString(441, 360, JUST_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,360,'INASISTENCIAS '+nn.Curso.Nombre)
							JUST_IIIBIM=nn.Nota
							c.drawString(463, 360, JUST_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,360,'INASISTENCIAS '+nn.Curso.Nombre)
							JUST_IVBIM=nn.Nota
							c.drawString(485, 360, JUST_IVBIM)#Ubicado en la columan IV Bimestre
							if str(JUST_IBIM)!='' and str(JUST_IIBIM)!='' and str(JUST_IIIBIM)!='' and str(JUST_IVBIM)!='':
								JUST1 = str(JUST_IBIM)
								JUST2 = str(JUST_IIBIM)
								JUST3 = str(JUST_IIIBIM)
								JUST4 = str(JUST_IVBIM)
								PRO_JUST=asistencias(JUST1,JUST2,JUST3,JUST4)
								c.drawString(517, 360,cero_izq(PRO_JUST))

					if nombre_curso=='INJU':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,345,'INASISTENCIAS '+nn.Curso.Nombre)
							INJU_IBIM=nn.Nota
							c.drawString(419, 345,INJU_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,345,'INASISTENCIAS '+nn.Curso.Nombre)
							INJU_IIBIM=nn.Nota
							c.drawString(441, 345, INJU_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,345,'INASISTENCIAS '+nn.Curso.Nombre)
							INJU_IIIBIM=nn.Nota
							c.drawString(463, 345, INJU_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,345,'INASISTENCIAS '+nn.Curso.Nombre)
							INJU_IVBIM=nn.Nota
							c.drawString(485, 345, INJU_IVBIM)#Ubicado en la columan IV Bimestre
							if str(INJU_IBIM)!='' and str(INJU_IIBIM)!='' and str(INJU_IIIBIM)!='' and str(INJU_IVBIM)!='':
								INJU1 = str(INJU_IBIM)
								INJU2 = str(INJU_IIBIM)
								INJU3 = str(INJU_IIIBIM)
								INJU4 = str(INJU_IVBIM)
								PRO_INJU=asistencias(INJU1,INJU2,INJU3,INJU4)
								c.drawString(517, 345,cero_izq(PRO_INJU))
					if nombre_curso=='TARD':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,330,nn.Curso.Nombre)
							TARD_IBIM=nn.Nota
							c.drawString(419, 330,TARD_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,330,nn.Curso.Nombre)
							TARD_IIBIM=nn.Nota
							c.drawString(441, 330, TARD_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,330,nn.Curso.Nombre)
							TARD_IIIBIM=nn.Nota
							c.drawString(463, 330, TARD_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,330,nn.Curso.Nombre)
							TARD_IVBIM=nn.Nota
							c.drawString(485, 330, TARD_IVBIM)#Ubicado en la columan IV Bimestre
							if str(TARD_IBIM)!='' and str(TARD_IIBIM)!='' and str(TARD_IIIBIM)!='' and str(TARD_IVBIM)!='':
								TARD1 = str(TARD_IBIM)
								TARD2 = str(TARD_IIBIM)
								TARD3 = str(TARD_IIIBIM)
								TARD4 = str(TARD_IVBIM)
								PRO_TARD=asistencias(TARD1,TARD2,TARD3,TARD4)
								c.drawString(517, 330,cero_izq(PRO_TARD))
					if nombre_curso=='APRE':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							APRE_IBIM=nn.Nota
							c.drawString(155, 285 - bloq_apre,APRE_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							APRE_IIBIM=nn.Nota
							c.drawString(155, 270 - bloq_apre, APRE_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							APRE_IIIBIM=nn.Nota
							c.drawString(155, 255 - bloq_apre, APRE_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							APRE_IVBIM=nn.Nota
							c.drawString(155, 240 - bloq_apre, APRE_IVBIM)#Ubicado en la columan IV Bimestre
					
				#try:

				punto=FinalSecun2(PRO_GENERAL)
				if punto>3:
					#Repite
					#c.drawString(374, 159, "X") # se comento x k nadie repite 2021
					
					#Recuperacion se puso recuperacion temporalmente x k nadie repite
					c.drawString(374, 174, "X")
				if punto>0 and punto<4:
					#Recuperacion
					c.drawString(374, 174, "X")
						
				if punto==0:
					#Promovido
					c.drawString(374, 189, "X")
				
				PRO_GENERAL=[]
				#except:
					#pass
					#print("ErrorcitodePromedio")
			else:

				MAT4=''
				COM4=''
				PER4=''
				CYT4=''
				cant=0
				#LIBRETA PRIMARIA #LIBRETA PRIMARIA  #LIBRETA PRIMARIA #LIBRETA PRIMARIA
				#LIBRETA PRIMARIA #LIBRETA PRIMARIA  #LIBRETA PRIMARIA #LIBRETA PRIMARIA
				#######################################################################################################
				bloq_epf=-15
				lineas_epf=-45
				c.drawImage("static/img/firma_maria.png", 90, 30, width=130, height=50)#firma ruth
				c.drawImage("static/img/clave_evaluacion.png", 430, 375, width=130, height=60)#firma ruth
				c.rect(430,647,115,15)#Cuadro Situación Final
				c.drawString(450,650,"SITUACIÓN FINAL")#Texto Situacion Final
				c.rect(430,615,115,32)#Cuadro para PROMOVIDO, REPITE, PASA A RECUPERACION
				
				c.drawString(175, 630, "ÁREA")#Area
				c.drawString(360, 650, "BIMESTRES")#Bimestres
				c.drawString(344, 622, "I")#Bimestres
				c.drawString(366, 622, "II")#Bimestres
				c.drawString(386, 622, "III")#Bimestres
				c.drawString(407, 622, "IV")#Bimestres

				c.roundRect(85, h-230, 252, 50,0)#Rectangulo Izquierdo Area
				c.roundRect(337, h-230, 22, 35,0)#Rectangulo I Bimestre
				c.roundRect(359, h-230, 22, 35,0)#Rectangulo II Bimestre
				c.roundRect(381, h-230, 22, 35,0)#Rectangulo III Bimestre
				c.roundRect(403, h-230, 22, 35,0)#Rectangulo IV Bimestre
				c.roundRect(337, h-195, 88, 15,0)#Rectangulo Bimestres
				#Lineas de evaluacion actitudinal
				c.line(337,432,337,357)#x1 x2
				c.line(359,432,359,357)#x1 x2
				c.line(381,432,381,357)#x1 x2
				c.line(403,432,403,357)#x1 x2

				#Lineas dE bIMESTRES
				c.line(337,612,337,462)#x1 x2
				c.line(359,612,359,462)#x1 x2
				c.line(381,612,381,462)#x1 x2
				c.line(403,612,403,462)#x1 x2

				c.roundRect(85,597, 340, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,582, 340, 15,0)#Rectangulo Cada Curso
				c.roundRect(85,567, 340, 15,0)#Rectangulo Cada Curso
				c.roundRect(85,552, 340, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,537, 340, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,522, 340, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,507, 340, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,492, 340, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,477, 340, 15,0)#Rectangulo Cada Curso						
				c.roundRect(85,462, 340, 15,0)#Rectangulo Cada Curso						
				c.setFont('Helvetica-Bold', 9, leading=None)
				c.drawString(190, 436, "EVALUACIÓN ACTITUDINAL")				
				c.drawString(190, 331, "INASISTENCIAS")
				
				c.drawString(130, 256, "EVALUACIÓN DEL PADRE DE FAMILIA")
				#####
				c.drawString(140, 144, "APRECIACIÓN DEL TUTOR (A) SOBRE LAS ACTITUDES DEL ALUMNO(A)")
				c.drawString(90, 129, "I BIMESTRE")
				c.drawString(90, 114, "II BIMESTRE")
				c.drawString(90, 99, "III BIMESTRE")
				c.drawString(90, 84, "IV BIMESTRE")
				c.line(150,141,150,81)#x1 x2
				#########
				c.setFont('Helvetica', 9, leading=None)
				
				c.roundRect(85,432, 340, 15,0)#EVALUACION ACTITUDINAL
				c.roundRect(85,417, 340, 15,0)#RESPETO
				c.roundRect(85,402, 340, 15,0)#PUNTUALIDAD
				c.roundRect(85,387, 340, 15,0)#ORDEN
				c.roundRect(85,372, 340, 15,0)#DISCIPLINA
				c.roundRect(85,357, 340, 15,0)#PARTICIPACIÓN

				c.roundRect(85,252, 340, 15,0)#EVALUACION PADRE DE FAMILIA TITULO
				c.roundRect(85,237, 340, 15,0)#1
				c.roundRect(85,222, 340, 15,0)#2
				c.roundRect(85,207, 340, 15,0)#3
				#c.roundRect(85,192, 340, 15,0)#4
				#c.roundRect(85,177, 340, 15,0)#5
				#c.roundRect(85,162, 340, 15,0)#6
				


				c.line(337,327,337,282)#ASISTENCIAS IBIM
				c.line(359,327,359,282)#ASISTENCIAS IIBIM
				c.line(381,327,381,282)#ASISTENCIAS IIIBIM
				c.line(403,327,403,282)#ASISTENCIAS IVBIM

				c.roundRect(85,327, 340, 15,0)#ASISTENCIA
				c.roundRect(85,312, 340, 15,0)#Bimestre
				c.roundRect(85,297, 340, 15,0)# Bimestre 312,297,282,267
				c.roundRect(85,282, 340, 15,0)# Bimestre
				
				# APRECIACION DEL TUTOR
				c.roundRect(85,141, 440, 15,0)# I BIMESTRE
				c.roundRect(85,126, 440, 15,0)# I BIMESTRE
				c.roundRect(85,111, 440, 15,0)# IIBIMESTRE
				c.roundRect(85,96, 440, 15,0)# III BIMESTRE
				c.roundRect(85, 81, 440, 15,0)# IV BIMESTRE

				##Lineas de Apreciacion
				c.line(337,252,337,162-lineas_epf)#Apreciacion IBIM
				c.line(359,252,359,162-lineas_epf)#Apreciacion IIBIM
				c.line(381,252,381,162-lineas_epf)#Apreciacion IIIBIM
				c.line(403,252,403,162-lineas_epf)#Apreciacion IVBIM

			#Desde aqui pasamos el for para generar cada rectangulo
				fila=h-230
				
				for nn in noti:
					fila = fila - 15
					nombre_curso=str(nn.Curso.Nombre)[0:4]#extra las 3 letras primera	
					if nombre_curso=='MATE':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,600,nn.Curso.Nombre)
							MATE_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(MATE_IBIM)))
							c.drawString(344,600, MATE_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,600,nn.Curso.Nombre)
							MATE_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(MATE_IIBIM)))
							c.drawString(366,600,MATE_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,600,nn.Curso.Nombre)
							MATE_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(MATE_IIIBIM)))
							c.drawString(386,600,MATE_IIIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,600,nn.Curso.Nombre)
							MATE_IVBIM=nn.Nota
							MAT4=MATE_IVBIM
							c.setFillColor(HexColor(letra_color(MATE_IVBIM)))
							c.drawString(407,600,MATE_IVBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
					if nombre_curso=='COMU':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,585,nn.Curso.Nombre)
							COMU_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(COMU_IBIM)))
							c.drawString(344, 585,COMU_IBIM)
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,585,nn.Curso.Nombre)
							COMU_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(COMU_IIBIM)))
							c.drawString(366, 585, COMU_IIBIM)#Ubicado en la columan II Bimestre		
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,585,nn.Curso.Nombre)
							COMU_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(COMU_IIIBIM)))
							c.drawString(386, 585, COMU_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,585,nn.Curso.Nombre)
							COMU_IVBIM=nn.Nota
							COM4=COMU_IVBIM
							c.setFillColor(HexColor(letra_color(COMU_IVBIM)))
							c.drawString(407, 585, COMU_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
					if str(nn.Curso.Nombre).find("ARTE")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,570,nn.Curso.Nombre)
							ARTE_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(ARTE_IBIM)))
							c.drawString(344, 570,ARTE_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,570,nn.Curso.Nombre)
							ARTE_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(ARTE_IIBIM)))
							c.drawString(366, 570, ARTE_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,570,nn.Curso.Nombre)
							ARTE_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(ARTE_IIIBIM)))
							c.drawString(386, 570, ARTE_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,570,nn.Curso.Nombre)
							ARTE_IVBIM=nn.Nota
							
							c.setFillColor(HexColor(letra_color(ARTE_IVBIM)))
							c.drawString(407, 570, ARTE_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("PERSONAL SOCIA")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,555,nn.Curso.Nombre)
							PERS_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PERS_IBIM)))	
							c.drawString(344, 555,PERS_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,555,nn.Curso.Nombre)
							PERS_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PERS_IIBIM)))	
							c.drawString(366, 555, PERS_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,555,nn.Curso.Nombre)
							PERS_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PERS_IIIBIM)))	
							c.drawString(386, 555, PERS_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,555,nn.Curso.Nombre)
							PERS_IVBIM=nn.Nota
							PER4=PERS_IVBIM
							c.setFillColor(HexColor(letra_color(PERS_IVBIM)))	
							c.drawString(407, 555, PERS_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("FÍSI")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,540,nn.Curso.Nombre)
							EFIS_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(EFIS_IBIM)))	
							c.drawString(344, 540,EFIS_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,540,nn.Curso.Nombre)
							EFIS_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(EFIS_IIBIM)))	
							c.drawString(366, 540, EFIS_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,540,nn.Curso.Nombre)
							EFIS_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(EFIS_IIIBIM)))	
							c.drawString(386, 540, EFIS_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,540,nn.Curso.Nombre)
							EFIS_IVBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(EFIS_IVBIM)))	
							c.drawString(407, 540, EFIS_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("RELI")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,525,nn.Curso.Nombre)
							RELI_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(RELI_IBIM)))	
							c.drawString(344, 525,RELI_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,525,nn.Curso.Nombre)
							RELI_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(RELI_IIBIM)))
							c.drawString(366, 525, RELI_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,525,nn.Curso.Nombre)
							RELI_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(RELI_IIIBIM)))
							c.drawString(386, 525, RELI_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,525,nn.Curso.Nombre)
							RELI_IVBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(RELI_IVBIM)))
							c.drawString(407, 525, RELI_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("CIENCIA Y TECNOLOG")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,510,nn.Curso.Nombre)
							CIEN_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(CIEN_IBIM)))
							c.drawString(344, 510,CIEN_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,510,nn.Curso.Nombre)
							CIEN_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(CIEN_IIBIM)))
							c.drawString(366, 510, CIEN_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,510,nn.Curso.Nombre)
							CIEN_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(CIEN_IIIBIM)))
							c.drawString(386, 510, CIEN_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,510,nn.Curso.Nombre)
							CIEN_IVBIM=nn.Nota
							CYT4=CIEN_IVBIM
							c.setFillColor(HexColor(letra_color(CIEN_IVBIM)))
							c.drawString(407, 510, CIEN_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))

					if nombre_curso=='INGL':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,495,nn.Curso.Nombre)
							INGL_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(INGL_IBIM)))
							c.drawString(344, 495,INGL_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,495,nn.Curso.Nombre)
							INGL_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(INGL_IIBIM)))
							c.drawString(366, 495, INGL_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,495,nn.Curso.Nombre)
							INGL_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(INGL_IIIBIM)))
							c.drawString(386, 495, INGL_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,495,nn.Curso.Nombre)
							INGL_IVBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(INGL_IVBIM)))
							c.drawString(407, 495, INGL_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
					
					if nombre_curso=='CHIN':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,480,nn.Curso.Nombre)
							CHIN_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(CHIN_IBIM)))
							c.drawString(344, 480,CHIN_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,480,nn.Curso.Nombre)
							CHIN_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(CHIN_IIBIM)))
							c.drawString(366, 480, CHIN_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,480,nn.Curso.Nombre)
							CHIN_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(CHIN_IIIBIM)))
							c.drawString(386, 480, CHIN_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,480,nn.Curso.Nombre)
							CHIN_IVBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(CHIN_IVBIM)))
							c.drawString(407, 480, CHIN_IVBIM)
							c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("COMPUTAC")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,465,nn.Curso.Nombre)
							PUTA_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PUTA_IBIM)))
							c.drawString(344, 465,PUTA_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,465,nn.Curso.Nombre)
							PUTA_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PUTA_IIBIM)))
							c.drawString(366, 465, PUTA_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,465,nn.Curso.Nombre)
							PUTA_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PUTA_IIIBIM)))
							c.drawString(386, 465, PUTA_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,465,nn.Curso.Nombre)
							PUTA_IVBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PUTA_IVBIM)))
							c.drawString(407, 465, PUTA_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
			

					if str(nn.Curso.Nombre).find("RESPETO")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,421,nn.Curso.Nombre)
							RESP_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(RESP_IBIM)))
							c.drawString(344, 421,RESP_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,421,nn.Curso.Nombre)
							RESP_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(RESP_IIBIM)))
							c.drawString(366, 421, RESP_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,421,nn.Curso.Nombre)
							RESP_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(RESP_IIIBIM)))
							c.drawString(386, 421, RESP_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,421,nn.Curso.Nombre)
							RESP_IVBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(RESP_IVBIM)))
							c.drawString(407, 421, RESP_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("PUNTUALIDAD")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,406,nn.Curso.Nombre)
							PUNT_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PUNT_IBIM)))
							c.drawString(344, 406,PUNT_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,406,nn.Curso.Nombre)
							PUNT_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PUNT_IIBIM)))
							c.drawString(366, 406, PUNT_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,406,nn.Curso.Nombre)
							PUNT_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PUNT_IIIBIM)))
							c.drawString(386, 406, PUNT_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,406,nn.Curso.Nombre)
							PUNT_IVBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PUNT_IVBIM)))
							c.drawString(407, 406, PUNT_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))

					if nombre_curso=='ORDE':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,391,nn.Curso.Nombre)
							ORD_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(ORD_IBIM)))
							c.drawString(344, 391,ORD_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,391,nn.Curso.Nombre)
							ORD_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(ORD_IIBIM)))
							c.drawString(366, 391, ORD_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,391,nn.Curso.Nombre)
							ORD_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(ORD_IIIBIM)))
							c.drawString(386, 391, ORD_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,391,nn.Curso.Nombre)
							ORD_IVBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(ORD_IVBIM)))
							c.drawString(407, 391, ORD_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("DISCIPLINA")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,376,nn.Curso.Nombre)
							DIS_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(DIS_IBIM)))
							c.drawString(344, 376,DIS_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,376,nn.Curso.Nombre)
							DIS_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(DIS_IIBIM)))
							c.drawString(366, 376, DIS_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,376,nn.Curso.Nombre)
							DIS_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(DIS_IIIBIM)))
							c.drawString(386, 376, DIS_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,376,nn.Curso.Nombre)
							DIS_IVBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(DIS_IVBIM)))
							c.drawString(407, 376, DIS_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("PARTICIPACIÓN")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,361,nn.Curso.Nombre)
							PAR_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PAR_IBIM)))
							c.drawString(344, 361,PAR_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,361,nn.Curso.Nombre)
							PAR_IIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PAR_IIBIM)))
							c.drawString(366, 361, PAR_IIBIM)#Ubicado en la columan II Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,361,nn.Curso.Nombre)
							PAR_IIIBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PAR_IIIBIM)))
							c.drawString(386, 361, PAR_IIIBIM)#Ubicado en la columan III Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,361,nn.Curso.Nombre)
							PAR_IVBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PAR_IVBIM)))
							c.drawString(407, 361, PAR_IVBIM)#Ubicado en la columan IV Bimestre
							c.setFillColor(HexColor(0x0A0A0A))
###########################################################################
					if nombre_curso=='JUST':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,315,'INASISTENCIAS '+nn.Curso.Nombre)
							JUST_IBIM=nn.Nota
							c.drawString(344, 315,JUST_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,315,'INASISTENCIAS '+nn.Curso.Nombre)
							JUST_IIBIM=nn.Nota
							c.drawString(366, 315, JUST_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,315,'INASISTENCIAS '+nn.Curso.Nombre)
							JUST_IIIBIM=nn.Nota
							c.drawString(386, 315, JUST_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,315,'INASISTENCIAS '+nn.Curso.Nombre)
							JUST_IVBIM=nn.Nota
							c.drawString(407, 315, JUST_IVBIM)#Ubicado en la columan IV Bimestre

					if nombre_curso=='INJU':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,300,'INASISTENCIAS '+nn.Curso.Nombre)
							INJU_IBIM=nn.Nota
							c.drawString(344, 300,INJU_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,300,'INASISTENCIAS '+nn.Curso.Nombre)
							INJU_IIBIM=nn.Nota
							c.drawString(366, 300, INJU_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,300,'INASISTENCIAS '+nn.Curso.Nombre)
							INJU_IIIBIM=nn.Nota
							c.drawString(386, 300, INJU_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,300,'INASISTENCIAS '+nn.Curso.Nombre)
							INJU_IVBIM=nn.Nota
							c.drawString(407, 300, INJU_IVBIM)#Ubicado en la columan IV Bimestre

					if nombre_curso=='TARD':
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,285,nn.Curso.Nombre)
							TARD_IBIM=nn.Nota
							c.drawString(344, 285,TARD_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,285,nn.Curso.Nombre)
							TARD_IIBIM=nn.Nota
							c.drawString(366, 285, TARD_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,285,nn.Curso.Nombre)
							TARD_IIIBIM=nn.Nota
							c.drawString(386, 285, TARD_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,285,nn.Curso.Nombre)
							TARD_IVBIM=nn.Nota
							c.drawString(407, 285, TARD_IVBIM)#Ubicado en la columan IV Bimestre

					if str(nn.Curso.Nombre).find("Asiste y participa en las clases")!= -1: ##Se cambión "Cumple Puntualemente"
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,240,nn.Curso.Nombre)
							PADR_IBIM=nn.Nota
							c.drawString(343, 240,PADR_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,240,nn.Curso.Nombre)
							PADR_IIBIM=nn.Nota
							c.drawString(366, 240, PADR_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,240,nn.Curso.Nombre)
							PADR_IIIBIM=nn.Nota
							c.drawString(386, 240, PADR_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,240,nn.Curso.Nombre)
							PADR_IVBIM=nn.Nota
							c.drawString(407, 240, PADR_IVBIM)#Ubicado
			##		
					if str(nn.Curso.Nombre).find("Envía correctamente uniformado")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,225,nn.Curso.Nombre)
							PADR2_IBIM=nn.Nota
							c.drawString(343, 225,PADR2_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,225,nn.Curso.Nombre)
							PADR2_IIBIM=nn.Nota
							c.drawString(366, 225, PADR2_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,225,nn.Curso.Nombre)
							PADR2_IIIBIM=nn.Nota
							c.drawString(386, 225, PADR2_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,225,nn.Curso.Nombre)
							PADR2_IVBIM=nn.Nota
							c.drawString(407, 225, PADR2_IVBIM)#Ubicado

					if str(nn.Curso.Nombre).find("Justifica la inasistencia de su hi")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,210-bloq_epf,nn.Curso.Nombre)
							PADR3_IBIM=nn.Nota
							c.drawString(343, 210-bloq_epf,PADR3_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,210-bloq_epf,nn.Curso.Nombre)
							PADR3_IIBIM=nn.Nota
							c.drawString(366, 210-bloq_epf, PADR3_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,210-bloq_epf,nn.Curso.Nombre)
							PADR3_IIIBIM=nn.Nota
							c.drawString(386, 210-bloq_epf, PADR3_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,210-bloq_epf,nn.Curso.Nombre)
							PADR3_IVBIM=nn.Nota
							c.drawString(407, 210-bloq_epf, PADR3_IVBIM)#Ubicado

					if str(nn.Curso.Nombre).find("Apoya en las tareas asignad")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,195-bloq_epf,nn.Curso.Nombre)
							PADR4_IBIM=nn.Nota
							c.drawString(343, 195-bloq_epf,PADR4_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,195-bloq_epf,nn.Curso.Nombre)
							PADR4_IIBIM=nn.Nota
							c.drawString(366, 195-bloq_epf, PADR4_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,195-bloq_epf,nn.Curso.Nombre)
							PADR4_IIIBIM=nn.Nota
							c.drawString(386, 195-bloq_epf, PADR4_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,195-bloq_epf,nn.Curso.Nombre)
							PADR4_IVBIM=nn.Nota
							c.drawString(407, 195-bloq_epf, PADR4_IVBIM)#Ubicado

					if str(nn.Curso.Nombre).find("Asiste a reuniones programa")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,180,nn.Curso.Nombre)
							PADR5_IBIM=nn.Nota
							c.drawString(343, 180,PADR5_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,180,nn.Curso.Nombre)
							PADR5_IIBIM=nn.Nota
							c.drawString(366, 180, PADR5_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,180,nn.Curso.Nombre)
							PADR5_IIIBIM=nn.Nota
							c.drawString(386, 180, PADR5_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,180,nn.Curso.Nombre)
							PADR5_IVBIM=nn.Nota
							c.drawString(407, 180, PADR5_IVBIM)#Ubicado

					if str(nn.Curso.Nombre).find("Colabora con las necesidades")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							c.drawString(90,165,nn.Curso.Nombre)
							PADR6_IBIM=nn.Nota
							c.drawString(343, 165,PADR6_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							c.drawString(90,165,nn.Curso.Nombre)
							PADR6_IIBIM=nn.Nota
							c.drawString(366, 165, PADR6_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							c.drawString(90,165,nn.Curso.Nombre)
							PADR6_IIIBIM=nn.Nota
							c.drawString(386, 165, PADR6_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							c.drawString(90,165,nn.Curso.Nombre)
							PADR6_IVBIM=nn.Nota
							c.drawString(407, 165, PADR6_IVBIM)#Ubicado

					if str(nn.Curso.Nombre).find("PRECIACI")!= -1:
						if nn.PAcademico.Nombre=='1 BIMESTRE' or nn.PAcademico.Nombre=='I BIMESTRE':#
							APRE_IBIM=nn.Nota
							c.drawString(155, 130,APRE_IBIM)#ubicado en la columna IBimestre
						if nn.PAcademico.Nombre=='2 BIMESTRE' or nn.PAcademico.Nombre=='II BIMESTRE':#
							APRE2_IIBIM=nn.Nota
							c.drawString(155, 115, APRE2_IIBIM)#Ubicado en la columan II Bimestre
						if nn.PAcademico.Nombre=='3 BIMESTRE' or nn.PAcademico.Nombre=='III BIMESTRE':#
							APRE3_IIIBIM=nn.Nota
							c.drawString(155, 100, APRE3_IIIBIM)#Ubicado en la columan III Bimestre
						if nn.PAcademico.Nombre=='4 BIMESTRE' or nn.PAcademico.Nombre=='IV BIMESTRE':#
							APRE4_IVBIM=nn.Nota
							c.drawString(155, 85, APRE4_IVBIM)#Ubicado
							#try:
								#ENCONTRADO SITUACION FINAL DE PRIMARIA#######
								#if MATE_IVBIM != '' and COMU_IVBIM != '' and PERS_IVBIM != '' and CIEN_IVBIM != '':
					if (nn.PAcademico.Nombre).find('4 BIMESTRE')!= -1 or (nn.PAcademico.Nombre).find('IV BIMESTRE')!= -1:				
						if str(m.Grado)=='1PRIM':
							c.drawString(450,630, "PROMOVIDO")#Ubicado
						else:
							c.drawString(450,630, FinalPrim2(m.Grado,MAT4,COM4,PER4,CYT4))
							
			c.showPage()

		c.save()#Guarda el documento fuera del bucle
		pdf = buffer.getvalue()
		buffer.close()
		response.write(pdf)
		
		return response
			#return render(request,'otras_opciones/base_reportes.html',contexto)
	else:
		return render(request,'otras_opciones/imprimir_libretas.html',contexto)


######AVANCES########AVANCES########AVANCES########AVANCES########AVANCES########AVANCES
########AVANCES########AVANCES########AVANCES########AVANCES########AVANCES########AVANCES
########AVANCES########AVANCES########AVANCES########AVANCES########AVANCES########AVANCES
def ImprimirAvances(request):
	#matri = Matricula() comentado 16-Jul-2019 10:25 p.m.
	
	#########variables para mover los bloques del avance
	bloq_cursos_avance=60
	bloq_datos_avance=35
	
	bloq_firma_avance_prim=0
	bloq_firma_avance_sec=70
	#################################################

	aac = AnoAcademico.objects.all().order_by('-Ano')
	pac = PAcademico.objects.all().order_by('Nombre')
	contexto = {'aac':aac,'pac':pac}
	if request.method == 'POST':
		paca = PAcademico()
		paca.id = request.POST.get("Pac")
		
		ano = AnoAcademico() 
		ano.id = request.POST.get("Ano")

		gra = request.POST.get("Grado")
		sec = request.POST.get("Seccion")
		mat = Matricula.objects.filter(Grado=gra,Seccion=sec,AnoAcademico=ano).order_by('Alumno__ApellidoPaterno','Alumno__ApellidoMaterno','Alumno__Nombres')
		pdf_file_name=gra+'_'+sec+'.pdf'#nombre del archivo
		
		tutor = Docente.objects.get(TutorGrado=gra,TutorSeccion=sec)#Obtener Tutor	
		response = HttpResponse(content_type='application/pdf')
		response['Content-Disposition'] = 'inline; filename="'+pdf_file_name+'"'#del ejemplo

		buffer = BytesIO()#Del ejemplo
		c = canvas.Canvas(buffer, pagesize=portrait(A4))
		paca = PAcademico.objects.get(id=paca.id)#Bimestre
		for m in mat:#Para empezar a dibujar una libreta por cada matricula es una libreta
			c.drawImage("static/img/logo_cvallejo.png", 95, 715, width=55, height=70)
			c.drawImage("static/img/logo_goreloreto.png", 105, 790, width=80, height=35)
			c.drawImage("static/img/logo_minedu.png", 280, 780, width=75, height=45)
			c.drawImage("static/img/logo_dreloreto.png", 450, 780, width=50, height=50)
			
			c.setFont('Helvetica-Bold', 12, leading=None)
			c.setFillColor(HexColor(0x4C9141))
			c.drawCentredString(300, 765, "COLEGIO COOPERATIVO CESAR VALLEJO")
			c.setFont('Helvetica-Bold', 11, leading=None)
			c.setFillColor(HexColor(0x0A0A0A))
			c.drawCentredString(300, 755, "NIVEL PRIMARIO-SECUNDARIO")
			c.setFont('Helvetica', 10, leading=None)	
			c.drawCentredString(300, 745, "Putumayo N° 966-Iquitos-Telef.973 891800")
			c.drawCentredString(300, 735, "Email: colegio_vallejo@yahoo.com")
			c.setFont('Helvetica-Bold', 12, leading=None)
			c.drawCentredString(300, 720, "AVANCE DE NOTAS " + str(m.AnoAcademico))
			
			c.setFont('Helvetica', 9, leading=None)			
			c.drawString(90, 705-bloq_datos_avance, "ALUMNO (A):")
			c.drawString(90, 690-bloq_datos_avance, "GRADO:")
			c.drawString(420, 705-bloq_datos_avance,"NIVEL:")
			c.drawString(420, 690-bloq_datos_avance,"BIMESTRE:")
			c.drawString(90, 675-bloq_datos_avance, "TUTOR (A):")
			
			c.drawString(142, 675-bloq_datos_avance, tutor.User.first_name + ' ' + tutor.User.last_name)

			w,h = A4#toma como referencia para hacer los rectangulos
			c.roundRect(85, 666-bloq_datos_avance,460,50,3)
			c.setFont('Helvetica',9, leading=None)
			grado=str(m.Grado)[0]#Extrae solo la primera letra
			c.drawString(130, 690-bloq_datos_avance, str(grado+'°'))
			c.drawString(145, 690-bloq_datos_avance, str('"'+m.Seccion+'"'))
			c.drawString(150, 705-bloq_datos_avance, str(m.Alumno.Nombres +' '+m.Alumno.ApellidoPaterno+' '+m.Alumno.ApellidoMaterno).upper())
			
			nivel=str(m.Grado)[1:len(m.Grado)] #extrae solo PRIM
			if nivel=='PRIM':
				nivel='PRIMARIO'
			else:
				nivel='SECUNDARIO'
			c.drawString(470, 690-bloq_datos_avance, str(paca.Nombre).upper())#Periodo academico

			c.drawString(453, 705-bloq_datos_avance,str(nivel).upper())
			noti = AvanceNotas.objects.filter(Matricula__id=m.id,PAcademico__id=paca.id)

			#AVANCE SECUNDARIA #AVANCE SECUNDARIA #AVANCE SECUNDARIA #AVANCE SECUNDARIA			
			#AVANCE SECUNDARIA #AVANCE SECUNDARIA #AVANCE SECUNDARIA #AVANCE SECUNDARIA			
			if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC' or str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
				c.drawImage("static/img/firma_ruth.png", 90, 150-bloq_firma_avance_sec, width=150, height=65)#firma ruth
				c.drawImage("static/img/firma_segundo.png", 350, 150-bloq_firma_avance_sec, width=150, height=70)#firma director encargado
				c.drawString(240, 630-bloq_cursos_avance, "ÁREA")#Area
				c.drawString(424, 628-bloq_cursos_avance, "PROMEDIO")#PROMEDIO
				
				if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':#
					c.drawString(497, 633-bloq_cursos_avance, "NOTA")#Bimestres
					c.drawString(486, 620-bloq_cursos_avance, "SIMULACRO")#
					c.roundRect(481, 611-bloq_cursos_avance, 64, 50,0)#Nota Simulacro
					###########287 x 337
					c.line(481,337-bloq_cursos_avance,481,612-bloq_cursos_avance)#Linea Vertical NotaSimulacro
					c.line(415,337-bloq_cursos_avance,415,612-bloq_cursos_avance)#x1 x2

					c.roundRect(85,587-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Curso
					c.roundRect(85,562-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Curso
					c.roundRect(85,537-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Curso						
					c.roundRect(85,512-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Curso						
					c.roundRect(85,487-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Curso						
					c.roundRect(85,462-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Cur
					c.roundRect(85,437-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Curso						
					c.roundRect(85,412-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Curso						
					c.roundRect(85,387-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Curso						
					c.roundRect(85,362-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Curso
					c.roundRect(85,337-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Curso
					#c.roundRect(85,312-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Curso
					#c.roundRect(85,287-bloq_cursos_avance, 460, 25,0)#Rectangulo Cada Curso
				else:						
					c.line(415,337-bloq_cursos_avance,415,612-bloq_cursos_avance)#x1 x2
					
					c.roundRect(85,587-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Curso
					c.roundRect(85,562-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Curso
					c.roundRect(85,537-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Curso						
					c.roundRect(85,512-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Curso						
					c.roundRect(85,487-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Curso						
					c.roundRect(85,462-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Cur
					c.roundRect(85,437-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Curso						
					c.roundRect(85,412-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Curso						
					c.roundRect(85,387-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Curso						
					c.roundRect(85,362-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Curso
					c.roundRect(85,337-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Curso
					#c.roundRect(85,312-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Curso
					#c.roundRect(85,287-bloq_cursos_avance, 396, 25,0)#Rectangulo Cada Curso

				c.roundRect(85, 611-bloq_cursos_avance, 330, 50,0)#Rectangulo Izquierdo Area
				c.roundRect(415, 611-bloq_cursos_avance, 66, 50,0)#Promedio	
				#Linea de vertical secundaria
				#c.line(503,447,503,612)#x1 x2								
			#Desde aqui pasamos el for para generar cada rectangulo
				fila=h-230
				for nn in noti:
					fila = fila - 15
					nombre_curso=str(nn.Curso.Nombre)[0:4]#extra las 3 letras primera	
					if nombre_curso=='MATE':
							c.drawString(90,595-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso Matematica
							MATE_IBIM=nn.Nota
							MATESIMU=nn.SimulacroNota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':# SE PONE ESTO PORQUE 1, 2 y3 SEC ES LETRA
								c.setFillColor(HexColor(pone_color(MATE_IBIM)))
								c.drawString(442,595-bloq_cursos_avance, numeros_a_letras(MATE_IBIM))#
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(pone_color(MATE_IBIM)))
								c.drawString(442,595-bloq_cursos_avance, MATE_IBIM)
								c.setFillColor(HexColor(0x0A0A0A))
								if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
									c.setFillColor(HexColor(pone_color(MATESIMU)))
									c.drawString(508,595-bloq_cursos_avance, MATESIMU)
									c.setFillColor(HexColor(0x0A0A0A))
					#RAZONAMIENTO MATEMÁTICO SERÁ ELIMINADO
					# if str(nn.Curso.Nombre).find("RAZONAMIENTO MAT")!= -1:
					# 		c.drawString(90,570-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
					# 		RAZM_IBIM=nn.Nota
					# 		RAZMSIMU=nn.SimulacroNota
					# 		if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
					# 			c.setFillColor(HexColor(pone_color(RAZM_IBIM)))
					# 			c.drawString(442, 570-bloq_cursos_avance,numeros_a_letras(RAZM_IBIM))#ubicado en la columna IBimestre
					# 			c.setFillColor(HexColor(0x0A0A0A))
					# 		else:
					# 			c.setFillColor(HexColor(pone_color(RAZM_IBIM)))
					# 			c.drawString(442, 570-bloq_cursos_avance,RAZM_IBIM)#ubicado en la columna IBimestre
					# 			c.setFillColor(HexColor(0x0A0A0A))
					# 			if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
					# 				c.setFillColor(HexColor(pone_color(RAZMSIMU)))
					# 				c.drawString(508, 570-bloq_cursos_avance,RAZMSIMU)#ubicado en la columna IBimestre
					# 				c.setFillColor(HexColor(0x0A0A0A))
					if nombre_curso=='COMU':
							c.drawString(90,570-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							COMU_IBIM=nn.Nota
							COMUSIMU=nn.SimulacroNota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								c.setFillColor(HexColor(pone_color(COMU_IBIM)))
								c.drawString(442, 570-bloq_cursos_avance,numeros_a_letras(COMU_IBIM))
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(pone_color(COMU_IBIM)))
								c.drawString(442, 570-bloq_cursos_avance,COMU_IBIM)
								c.setFillColor(HexColor(0x0A0A0A))
								if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
									c.setFillColor(HexColor(pone_color(COMUSIMU)))
									c.drawString(508, 570-bloq_cursos_avance,COMUSIMU)
									c.setFillColor(HexColor(0x0A0A0A))
					#RAZONAMIENTO VERBAL SERÁ ELIMINADO
					# if str(nn.Curso.Nombre).find("RAZONAMIENTO VER")!= -1:
					# 		c.drawString(90,520-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
					# 		COMU_IBIM=nn.Nota
					# 		RAZVSIMU=nn.SimulacroNota
					# 		if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
					# 			c.setFillColor(HexColor(pone_color(COMU_IBIM)))
					# 			c.drawString(442, 520-bloq_cursos_avance,numeros_a_letras(COMU_IBIM))
					# 			c.setFillColor(HexColor(0x0A0A0A))
					# 		else:
					# 			c.setFillColor(HexColor(pone_color(COMU_IBIM)))
					# 			c.drawString(442, 520-bloq_cursos_avance,COMU_IBIM)
					# 			c.setFillColor(HexColor(0x0A0A0A))
					# 			if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
					# 				c.setFillColor(HexColor(pone_color(RAZVSIMU)))
					# 				c.drawString(508, 520-bloq_cursos_avance,RAZVSIMU)
					# 				c.setFillColor(HexColor(0x0A0A0A))

					if nombre_curso=='INGL':
							c.drawString(90,545-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							INGL_IBIM=nn.Nota
							INGLSIMU=nn.SimulacroNota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								c.setFillColor(HexColor(pone_color(INGL_IBIM)))
								c.drawString(442, 545-bloq_cursos_avance,numeros_a_letras(INGL_IBIM))
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(pone_color(INGL_IBIM)))
								c.drawString(442, 545-bloq_cursos_avance,INGL_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
								if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
									c.setFillColor(HexColor(pone_color(INGLSIMU)))
									c.drawString(508, 545-bloq_cursos_avance,INGLSIMU)#ubicado en la columna IBimestre
									c.setFillColor(HexColor(0x0A0A0A))
					
					if nombre_curso=='CHIN':
							c.drawString(90,520-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							CHIN_IBIM=nn.Nota
							CHINSIMU=nn.SimulacroNota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								c.setFillColor(HexColor(pone_color(CHIN_IBIM)))
								c.drawString(442, 520-bloq_cursos_avance,numeros_a_letras(CHIN_IBIM))
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(pone_color(CHIN_IBIM)))
								c.drawString(442, 520-bloq_cursos_avance,CHIN_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
								if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
									c.setFillColor(HexColor(pone_color(CHINSIMU)))
									c.drawString(508, 520-bloq_cursos_avance,CHINSIMU)#ubicado en la columna IBimestre
									c.setFillColor(HexColor(0x0A0A0A))
					if nombre_curso=='ARTE':
							c.drawString(90,495-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							ARTE_IBIM=nn.Nota
							ARTESIMU=nn.SimulacroNota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								c.setFillColor(HexColor(pone_color(ARTE_IBIM)))
								c.drawString(442, 495-bloq_cursos_avance,numeros_a_letras(ARTE_IBIM))
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(pone_color(ARTE_IBIM)))
								c.drawString(442, 495-bloq_cursos_avance,ARTE_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
								if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
									c.setFillColor(HexColor(pone_color(ARTESIMU)))
									c.drawString(508, 495-bloq_cursos_avance,ARTESIMU)#
									c.setFillColor(HexColor(0x0A0A0A))
					if str(nn.Curso.Nombre).find("CIENCIAS SOCIALES")!= -1:
							c.drawString(90,470-bloq_cursos_avance,nn.Curso.Nombre)#
							CIES_IBIM=nn.Nota
							CIESSIMU=nn.SimulacroNota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								c.setFillColor(HexColor(pone_color(CIES_IBIM)))
								c.drawString(442, 470-bloq_cursos_avance,numeros_a_letras(CIES_IBIM))#
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(pone_color(CIES_IBIM)))
								c.drawString(442, 470-bloq_cursos_avance,CIES_IBIM)#
								c.setFillColor(HexColor(0x0A0A0A))
								if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
									c.setFillColor(HexColor(pone_color(CIESSIMU)))
									c.drawString(508, 470-bloq_cursos_avance,CIESSIMU)#
									c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("DESARROLLO PERSONAL")!= -1:
							c.drawString(90,445-bloq_cursos_avance,nn.Curso.Nombre)#
							DPER_IBIM=nn.Nota
							DPERSIMU=nn.SimulacroNota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								c.setFillColor(HexColor(pone_color(DPER_IBIM)))
								c.drawString(442, 445-bloq_cursos_avance,numeros_a_letras(DPER_IBIM))#
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(pone_color(DPER_IBIM)))	
								c.drawString(442, 445-bloq_cursos_avance,DPER_IBIM)#
								c.setFillColor(HexColor(0x0A0A0A))
								if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
									c.setFillColor(HexColor(pone_color(DPERSIMU)))	
									c.drawString(508, 445-bloq_cursos_avance,DPERSIMU)#
									c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("FÍSICA")!= -1:
							c.drawString(90,420-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							EFIS_IBIM=nn.Nota
							EFISSIMU=nn.SimulacroNota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								c.setFillColor(HexColor(pone_color(EFIS_IBIM)))
								c.drawString(442, 420-bloq_cursos_avance,numeros_a_letras(EFIS_IBIM))
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(pone_color(EFIS_IBIM)))	
								c.drawString(442, 420-bloq_cursos_avance,EFIS_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
								if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
									c.setFillColor(HexColor(pone_color(EFISSIMU)))	
									c.drawString(508, 420-bloq_cursos_avance,EFISSIMU)#ubicado en la columna IBimestre
									c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("RELIGI")!= -1:
							c.drawString(90,395-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							RELI_IBIM=nn.Nota
							RELISIMU=nn.SimulacroNota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								c.setFillColor(HexColor(pone_color(RELI_IBIM)))
								c.drawString(442, 395-bloq_cursos_avance,numeros_a_letras(RELI_IBIM))
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(pone_color(RELI_IBIM)))	
								c.drawString(442, 395-bloq_cursos_avance,RELI_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
								if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
									c.setFillColor(HexColor(pone_color(RELISIMU)))	
									c.drawString(508, 395-bloq_cursos_avance,RELISIMU)#ubicado en la columna IBimestre
									c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("CIENCIA Y TECNOLOG")!= -1:
							c.drawString(90,370-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							CIEN_IBIM=nn.Nota
							CIENSIMU=nn.SimulacroNota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								c.setFillColor(HexColor(pone_color(CIEN_IBIM)))
								c.drawString(442, 370-bloq_cursos_avance,numeros_a_letras(CIEN_IBIM))
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(pone_color(CIEN_IBIM)))
								c.drawString(442, 370-bloq_cursos_avance,CIEN_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
								if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
									c.setFillColor(HexColor(pone_color(CIENSIMU)))
									c.drawString(508, 370-bloq_cursos_avance,CIENSIMU)#ubicado en la columna IBimestre
									c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("TRABAJO")!= -1:
							c.drawString(90,345-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							TRAB_IBIM=nn.Nota
							TRABSIMU=nn.SimulacroNota
							if str(m.Grado)=='1SEC' or str(m.Grado)=='2SEC' or str(m.Grado)=='3SEC':
								c.setFillColor(HexColor(pone_color(TRAB_IBIM)))
								c.drawString(442, 345-bloq_cursos_avance,numeros_a_letras(TRAB_IBIM))
								c.setFillColor(HexColor(0x0A0A0A))
							else:
								c.setFillColor(HexColor(pone_color(TRAB_IBIM)))
								c.drawString(442, 345-bloq_cursos_avance,TRAB_IBIM)#ubicado en la columna IBimestre
								c.setFillColor(HexColor(0x0A0A0A))
								if str(m.Grado)=='4SEC' or str(m.Grado)=='5SEC':
									c.setFillColor(HexColor(pone_color(TRABSIMU)))
									c.drawString(508, 345-bloq_cursos_avance,TRABSIMU)#ubicado en la columna IBimestre
									c.setFillColor(HexColor(0x0A0A0A))
			else:#AVANCE PRIMARIA #AVANCE PRIMARIA #AVANCE PRIMARIA
				#######################################################################################################
				c.drawImage("static/img/firma_maria.png", 90, 110-bloq_firma_avance_prim, width=125, height=50)#firma ruth
				c.drawImage("static/img/firma_segundo.png", 350, 110-bloq_firma_avance_prim, width=150, height=70)#firma director encargado
				c.drawImage("static/img/clave_evaluacion.png", 435, 550-bloq_cursos_avance, width=130, height=60)#Clave Evaluacion
				c.setFont('Helvetica-Bold', 9, leading=None)
				
				#########comentado 24-nov-2020##################3
				#c.drawString(190, 322, "EVALUACIÓN ACTITUDINAL")
				###################################################
				
				c.drawString(175, 630-bloq_cursos_avance, "ÁREA")#Area
				c.drawString(355, 630-bloq_cursos_avance, "EVALUACIÓN")#Bimestres
				c.setFont('Helvetica', 9, leading=None)
				c.roundRect(85, 612-bloq_cursos_avance, 252, 50,0)#Rectangulo AREA
				
				c.roundRect(337, 612-bloq_cursos_avance, 88, 50,0)#Rectangulo EVALUACIÓN

				#Linea division curso de primaria
				c.line(337,612-bloq_cursos_avance,337,362-bloq_cursos_avance)#x1 x2

				#Lineas de evaluacion actitudinal######comentao 24-Nov-2020
				#c.line(337,312,337,187)#x1 x2 Evaluacion Actitudinal
				#####################################
				c.roundRect(85,587-bloq_cursos_avance, 340, 25,0)#Rectangulo Cada Curso
				c.roundRect(85,562-bloq_cursos_avance, 340, 25,0)#Rectangulo Cada Curso
				c.roundRect(85,537-bloq_cursos_avance, 340, 25,0)#Rectangulo Cada Curso						
				c.roundRect(85,512-bloq_cursos_avance, 340, 25,0)#Rectangulo Cada Curso						
				c.roundRect(85,487-bloq_cursos_avance, 340, 25,0)#Rectangulo Cada Curso						
				c.roundRect(85,462-bloq_cursos_avance, 340, 25,0)#Rectangulo Cada CurSO
				c.roundRect(85,437-bloq_cursos_avance, 340, 25,0)#Rectangulo Cada Curso						
				c.roundRect(85,412-bloq_cursos_avance, 340, 25,0)#Rectangulo Cada Curso						
				c.roundRect(85,387-bloq_cursos_avance, 340, 25,0)#Rectangulo Cada Curso						
				c.roundRect(85,362-bloq_cursos_avance, 340, 25,0)#Rectangulo Cada Curso

				###TCOMENTADO PARA 24-NOV-2020############3333
				#c.roundRect(85,312, 340, 25,0)#EVALUACION ACTITUDINAL
				#c.roundRect(85,287, 340, 25,0)#RESPETO
				#c.roundRect(85,262, 340, 25,0)#PUNTUALIDAD
				#c.roundRect(85,237, 340, 25,0)#ORDEN
				#c.roundRect(85,212, 340, 25,0)#DISCIPLINA
				#c.roundRect(85,187, 340, 25,0)#PARTICIPACIÓN
				####################################################33
				if str(m.Grado)=='1PRIM' or str(m.Grado)=='2PRIM' or str(m.Grado)=='3PRIM' or str(m.Grado)=='4PRIM':
					c.drawString(90,395-bloq_cursos_avance,"CHINO MANDARÍN")
					c.drawString(377, 395-bloq_cursos_avance,"-")
			#Desde aqui pasamos el for para generar cada rectangulo
				fila=h-230
				for nn in noti:
					fila = fila - 15
					nombre_curso=str(nn.Curso.Nombre)[0:4]#extra las 3 letras primera	
					if nombre_curso=='MATE':
							c.drawString(90,595-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo
							MATE_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(MATE_IBIM)))
							c.drawString(377,595-bloq_cursos_avance, MATE_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
					if nombre_curso=='COMU':
							c.drawString(90,570-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							COMU_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(COMU_IBIM)))
							c.drawString(377, 570-bloq_cursos_avance,COMU_IBIM)
							c.setFillColor(HexColor(0x0A0A0A))
					if str(nn.Curso.Nombre).find("ARTE")!= -1:
							c.drawString(90,545-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							ARTE_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(ARTE_IBIM)))
							c.drawString(377, 545-bloq_cursos_avance,ARTE_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
					if str(nn.Curso.Nombre).find("PERSONAL SOCIA")!= -1:
							c.drawString(90,520-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							PERS_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PERS_IBIM)))	
							c.drawString(377, 520-bloq_cursos_avance,PERS_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
					if str(nn.Curso.Nombre).find("FÍSI")!= -1:
							c.drawString(90,495-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							EFIS_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(EFIS_IBIM)))	
							c.drawString(377, 495-bloq_cursos_avance,EFIS_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
					if str(nn.Curso.Nombre).find("RELI")!= -1:
							c.drawString(90,470-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							RELI_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(RELI_IBIM)))	
							c.drawString(377, 470-bloq_cursos_avance,RELI_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
					if str(nn.Curso.Nombre).find("CIENCIA Y TECNOLOG")!= -1:
							c.drawString(90,445-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							CIEN_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(CIEN_IBIM)))
							c.drawString(377, 445-bloq_cursos_avance,CIEN_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
					if nombre_curso=='INGL':
							c.drawString(90,420-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							INGL_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(INGL_IBIM)))
							c.drawString(377, 420-bloq_cursos_avance,INGL_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))					
					if nombre_curso=='CHIN':
							c.drawString(90,395-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							CHIN_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(CHIN_IBIM)))
							c.drawString(377, 395-bloq_cursos_avance,CHIN_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))

					if str(nn.Curso.Nombre).find("COMPUTAC")!= -1:
							c.drawString(90,370-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
							PUTA_IBIM=nn.Nota
							c.setFillColor(HexColor(letra_color(PUTA_IBIM)))
							c.drawString(377, 370-bloq_cursos_avance,PUTA_IBIM)#ubicado en la columna IBimestre
							c.setFillColor(HexColor(0x0A0A0A))
					##todo esto es comentado porque pedido de selva 24-Nov-2020		
					# if str(nn.Curso.Nombre).find("RESPETO")!= -1:
					# 		c.drawString(90,295-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
					# 		RESP_IBIM=nn.Nota
					# 		c.setFillColor(HexColor(letra_color(RESP_IBIM)))
					# 		c.drawString(377, 295-bloq_cursos_avance,RESP_IBIM)#ubicado en la columna IBimestre
					# 		c.setFillColor(HexColor(0x0A0A0A))
					# if str(nn.Curso.Nombre).find("PUNTUALIDAD")!= -1:
					# 		c.drawString(90,270-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
					# 		PUNT_IBIM=nn.Nota
					# 		c.setFillColor(HexColor(letra_color(PUNT_IBIM)))
					# 		c.drawString(377, 270-bloq_cursos_avance,PUNT_IBIM)#ubicado en la columna IBimestre
					# 		c.setFillColor(HexColor(0x0A0A0A))
					# if nombre_curso=='ORDE':
					# 		c.drawString(90,245-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
					# 		ORD_IBIM=nn.Nota
					# 		c.setFillColor(HexColor(letra_color(ORD_IBIM)))
					# 		c.drawString(377, 245-bloq_cursos_avance,ORD_IBIM)#ubicado en la columna IBimestre
					# 		c.setFillColor(HexColor(0x0A0A0A))

					# if str(nn.Curso.Nombre).find("DISCIPLINA")!= -1:
					# 		c.drawString(90,220-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
					# 		DIS_IBIM=nn.Nota
					# 		c.setFillColor(HexColor(letra_color(DIS_IBIM)))
					# 		c.drawString(377, 220-bloq_cursos_avance,DIS_IBIM)#ubicado en la columna IBimestre
					# 		c.setFillColor(HexColor(0x0A0A0A))
					# if str(nn.Curso.Nombre).find("PARTICIPACIÓN")!= -1:
					# 		c.drawString(90,195-bloq_cursos_avance,nn.Curso.Nombre)#Escribiendo Curso COMUNICA
					# 		PAR_IBIM=nn.Nota
					# 		c.setFillColor(HexColor(letra_color(PAR_IBIM)))
					# 		c.drawString(377, 195-bloq_cursos_avance,PAR_IBIM)#ubicado en la columna IBimestre
					# 		c.setFillColor(HexColor(0x0A0A0A))
			c.showPage()
		c.save()#Guarda el documento fuera del bucle
		pdf = buffer.getvalue()
		buffer.close()
		response.write(pdf)
		
		return response
			#return render(request,'otras_opciones/base_reportes.html',contexto)
	else:
		return render(request,'otras_opciones/imprimir_avances.html',contexto)
def FinalPrim2(grado,mat,com,per,cyt):
	#num=0
	#mat=''
	#com=''
	#per=''
	#cyt=''

	#for x in notas:
	#	num=num+1
	#	if num==1:
	#		mat=x
	#	if num==2:
	#		com=x
	#	if num==3:
	#		per=x
	#	if num==4:
	#		cyt=x

	if str(mat) != '' and str(com)!='' and str(per)!='' and str(cyt)!='':
		if (grado=='2PRIM' or grado=='3PRIM' or grado=='4PRIM'):
			if (mat.find("A")!= -1 and com.find("A")!= -1):
				return ('PROMOVIDO')
			else:
				if (mat=='C' and com=='C'):
					return('RECUPERACIÓN')#AQUI DEBERIA SER REPITE SOLO 2021 NADIE REPITE X ESO SE PUSO ASI
				else:
					return('RECUPERACIÓN')
		else:
			#Arreglar Algoritmo # 5PRIM Y 6PRIM
			#if (grado=='5PRIM' or grado=='6PRIM'):
			if (mat.find("A")!= -1 and com.find("A")!= -1 and cyt.find("A")!= -1 and per.find("A")!= -1):
				return ('PROMOVIDO')
			else:
				if (mat=='C' and com=='C'):
					return('RECUPERACIÓN')#AQUI DEBERIA SER REPITE SOLO 2021 NADIE REPITE X ESO SE PUSO ASI
				else:
					return('RECUPERACIÓN')
	else:
		return('')
#	except:
	#return('--')


def FinalPrim(grado,mat,com,per,cyt):
	#El Curso CHINO MANDARIN no Aplica
	try:
		if (grado=='2PRIM' or grado=='3PRIM' or grado=='4PRIM'):
			if (mat.find("A")!= -1 and com.find("A")!= -1):
				return ('PROMOVIDO')
			else:
				if (mat=='C' and com=='C'):
					return('REPITE')
				else:
					return('RECUPERACIÓN')
		else:
			#Arreglar Algoritmo # 5PRIM Y 6PRIM
			#if (grado=='5PRIM' or grado=='6PRIM'):
			if (mat.find("A")!= -1 and com.find("A")!= -1 and cyt.find("A")!= -1 and per.find("A")!= -1):
				return ('PROMOVIDO')
			else:
				if (mat=='C' and com=='C'):
					return('REPITE')
				else:
					return('RECUPERACIÓN')
	except:
		return('ErrorNotaFinal')
def cero_izq(n):
	try:
		if int(n) < 10:
			valor="0" + str(n)
		else:
			valor=str(n)
		return valor
	except:
		return(str(n))

def asistencias(num1,num2,num3,num4):#PRIMARIA
	try:
		suma = (int(num1)+int(num2)+int(num3)+int(num4))
	except:
		suma = 0
	return suma

def calc_promedio(ns):
	try:
		cant=0
		suma=0
		notas = ns
		#Aqui empieza el bucle
		for x in notas:
			if IsNumber(int(x)):
				cant=cant+1
				suma=suma+int(x)
		prom=(round(float(int(suma)/cant)))
		return prom
	except:
		promedio='--'
		return promedio
		#Devuelve rayita si alguna de las notas no es número

def calculo_promedio(n1,n2,n3,n4):
	try:
		#promedio=(round(float(int(n1)+int(n2)+int(n3)+int(n4))/4))
		#return (promedio)
		cant=0
		suma=0
		notas = [n1,n2,n3,n4]
		#Aqui empieza el bucle
		for x in notas:
			if IsNumber(x):
				cant=cant+1
				suma=suma+int(x)
		prom=(round(float(int(suma)/cant)))
		return prom
	except:
		promedio='#'
		return promedio
		#Devuelve rayita si alguna de las notas no es número


def round(n):
	try:
		dec = n - int(n)
		rst = int(n)
		if dec < 0.5:
			rst = int(n)
		else:
			rst = int(n)+1
		return rst
	except:
		return ('-')#Porque algunos no tienen Nota Trasladados
    	
def IsNumber(numero):
	try:
		entero = int(numero)
		return True
	except:
		return False
def FinalSecun2(prom):
	pun=0
	datos=prom
	try:
		for x in datos:
			if	IsNumber(x):
				if int(x)<11:
					pun+=1
	except:	
		pun="--"
		return(pun)
	return pun
	
def FinalSecun(mat,com,ing,chi,art,cie,dpe,efi,rel,cta,tra):
	pun=0
	if IsNumber(mat):
		if mat<11:
			pun+=1
	if IsNumber(com):
		if com<11:
			pun+=1
	if IsNumber(ing):
		if ing<11:
			pun+=1
	if IsNumber(chi):
		if chi<11:
			pun+=1
	if IsNumber(art):
		if art<11:
			pun+=1
	if IsNumber(cie):
		if cie<11:
			pun+=1
	if IsNumber(dpe):
		if dpe<11:
			pun+=1
	if IsNumber(efi):
		if efi<11:
			pun+=1
	if IsNumber(rel):
		if rel<11:
			pun+=1
	if IsNumber(cta):
		if cta<11:
			pun+=1
	if IsNumber(tra):
		if tra<11:
			pun+=1
	return pun

def pone_color(valor):
	try:
		if int(valor) < 11:
			devuelve = '0xD62718'
		else:
			devuelve = '0x0A0A0A'
	except:
		devuelve = '0x0A0A0A'
		return devuelve
	return devuelve

def letra_color(valor):
	if str(valor)=='C':
		devuelve='0xD62718'#COLOR ROJO
	else:
		devuelve='0x0A0A0A'#COLOR NEGRO

	return devuelve

def numeros_a_letras(valor):
	try:
		if int(valor)>=0:
			if int(valor)<=10:
				devuelveLetra='C'
			else:
				if int(valor)<=13:
					devuelveLetra='B'
				else:
					if int(valor)<=16:
						devuelveLetra='A'
					else:
						if int(valor)<=20:
							devuelveLetra='AD'
						else:
							devuelveLetra=valor
	except:
		return valor
	return devuelveLetra
